from fastapi import FastAPI, Depends, Request, Form, Query, UploadFile, File, HTTPException
from fastapi.responses import RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from fastapi_login import LoginManager
from fastapi.staticfiles import StaticFiles
from passlib.context import CryptContext
from sqlalchemy import create_engine, Column, Integer, String, Numeric, DateTime, text, Boolean, ForeignKey
from sqlalchemy.orm import declarative_base, sessionmaker, relationship
from datetime import datetime, timedelta
from collections import Counter
import io
import csv
import os
import asyncio
from typing import Optional, List, Dict, Union, Any
from pydantic import BaseModel, validator, Field
from decimal import Decimal
from fastapi import WebSocket, WebSocketDisconnect

# Custom Jinja2 filters
def lt_filter(value, threshold):
    try:
        return float(value) < float(threshold)
    except (ValueError, TypeError):
        return False

def ternary_filter(value, true_val, false_val):
    return true_val if value else false_val

DATABASE_URL = "postgresql://tarificador_user:fr4v4t3l@localhost/tarificador"
SECRET = "secreto-super-importante"

engine = create_engine(DATABASE_URL)
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()
app = FastAPI()
# Montar archivos estáticos
app.mount("/static", StaticFiles(directory="static"), name="static")

templates = Jinja2Templates(directory="templates")
templates.env.filters['lt'] = lt_filter
templates.env.filters['ternary'] = ternary_filter
manager = LoginManager(SECRET, token_url="/auth/login", use_cookie=True)
manager.cookie_name = "auth_token"
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

# Add custom filters to Jinja2 environment
templates.env.filters['lt'] = lt_filter
templates.env.filters['ternary'] = ternary_filter

# Modelo para validar llamadas activas
class ActiveCall(BaseModel):
    callId: str
    origin: str
    destination: str
    startTime: str
    duration: int
    estimatedCost: float
    zone: Optional[str] = "Desconocida"

# Gestor de conexiones WebSocket
class ConnectionManager:
    def __init__(self):
        self.active_connections: List[WebSocket] = []
        self._lock = asyncio.Lock()  # Añadir un lock para evitar condiciones de carrera

    async def connect(self, websocket: WebSocket) -> None:
        await websocket.accept()
        async with self._lock:
            self.active_connections.append(websocket)

    def disconnect(self, websocket: WebSocket) -> None:
        if websocket in self.active_connections:
            self.active_connections.remove(websocket)

    async def broadcast(self, message: dict) -> None:
        closed_connections = []
        
        for connection in self.active_connections:
            try:
                await connection.send_json(message)
            except Exception:
                # Marcar la conexión para eliminarla después
                closed_connections.append(connection)
                
        # Eliminar las conexiones cerradas
        for conn in closed_connections:
            self.disconnect(conn)

ws_manager = ConnectionManager()

# Almacenamiento de llamadas activas con timestamp
active_calls: Dict[str, Dict] = {}
last_update = datetime.now()

# Estadísticas de WebSocket
ws_stats = {
    "total_connections": 0,
    "messages_sent": 0,
    "messages_received": 0
}

@app.get("/api/active-calls")
async def get_active_calls():
    """Obtiene la lista de llamadas activas para la API"""
    try:
        db = SessionLocal()
        
        active_calls_rows = db.execute(
            text("""SELECT call_id, calling_number, called_number, start_time, 
            current_duration, current_cost, zone
            FROM active_calls ORDER BY start_time DESC""")
        ).fetchall()
        
        active_calls = []
        for row in active_calls_rows:
            call = {
                "call_id": row[0],
                "calling_number": row[1],
                "called_number": row[2],
                "start_time": row[3].isoformat() if row[3] else None,
                "current_duration": row[4] if row[4] is not None else 0,
                "current_cost": float(row[5]) if row[5] is not None else 0.0,
                "zone": row[6] if len(row) > 6 else "Desconocida"
            }
            active_calls.append(call)
        
        return active_calls
    except Exception as e:
        print(f"Error obteniendo llamadas activas: {str(e)}")
        return {"status": "error", "message": str(e)}
        
@app.get("/api/active-calls-list")
def get_active_calls_list():
    db = SessionLocal()
    try:
        active_calls_rows = db.execute(
            text("SELECT * FROM active_calls ORDER BY start_time DESC")
        ).fetchall()
        
        active_calls = []
        for row in active_calls_rows:
            call = {
                "call_id": row[1],
                "calling_number": row[2],
                "called_number": row[3],
                "start_time": row[4].isoformat() if row[4] else None,
                "current_duration": row[6] if row[6] is not None else 0,
                "current_cost": float(row[7]) if row[7] is not None else 0.0
            }
            active_calls.append(call)
        
        return active_calls
    finally:
        db.close()

# Endpoint WebSocket principal
@app.websocket("/ws")
async def websocket_endpoint(websocket: WebSocket):
    await ws_manager.connect(websocket)
    print(f"Nueva conexión WebSocket establecida. Total conexiones: {len(ws_manager.active_connections)}")
    
    try:
        # Envía todas las llamadas activas al cliente que se acaba de conectar
        db = SessionLocal()
        active_calls_rows = db.execute(
            text("""SELECT call_id, calling_number, called_number, start_time, 
            current_duration, current_cost, connection_id 
            FROM active_calls""")
        ).fetchall()
        db.close()
        
        # Convertir resultados a lista de diccionarios con exactamente los campos esperados
        active_calls = []
        for row in active_calls_rows:
            call = {
                "call_id": row[0],
                "calling_number": row[1],
                "called_number": row[2],
                "start_time": row[3].isoformat() if row[3] else None,
                "current_duration": row[4] if row[4] is not None else 0,
                "current_cost": float(row[5]) if row[5] is not None else 0.0
            }
            active_calls.append(call)
        
        print(f"Enviando {len(active_calls)} llamadas activas iniciales: {active_calls}")
        
        # Asegurarse de que el formato sea exactamente lo que el cliente espera
        await websocket.send_json({
            "type": "update",
            "active_calls": active_calls
        })
        
        # Bucle principal para recibir mensajes del cliente
        while True:
            data = await websocket.receive_text()
            print(f"Mensaje recibido del cliente: {data}")
            
            try:
                message = json.loads(data)
                action = message.get("action")
                
                if action == "get_active_calls":
                    # Actualización manual solicitada por el cliente
                    db = SessionLocal()
                    active_calls_rows = db.execute(
                        text("""SELECT call_id, calling_number, called_number, start_time, 
                        current_duration, current_cost, connection_id 
                        FROM active_calls ORDER BY start_time DESC""")
                    ).fetchall()
                    db.close()
                    
                    # Convertir resultados a lista de diccionarios
                    active_calls = []
                    for row in active_calls_rows:
                        call = {
                            "call_id": row[0],
                            "calling_number": row[1],
                            "called_number": row[2],
                            "start_time": row[3].isoformat() if row[3] else None,
                            "current_duration": row[4] if row[4] is not None else 0,
                            "current_cost": float(row[5]) if row[5] is not None else 0.0,
                            "connection_id": row[6]
                        }
                        active_calls.append(call)
                    
                    await websocket.send_json({
                        "type": "update",
                        "active_calls": active_calls
                    })
                
                elif action == "terminate_call" and "call_id" in message:
                    # Procesar solicitud para terminar una llamada
                    call_id = message["call_id"]
                    print(f"Solicitud para terminar llamada: {call_id}")
                    
                    # Obtener connection_id de la llamada
                    db = SessionLocal()
                    result = db.execute(
                        text(f"SELECT connection_id FROM active_calls WHERE call_id = :call_id"),
                        {"call_id": call_id}
                    ).fetchone()
                    db.close()
                    
                    if result and result[0]:
                        # Implementar la terminación de la llamada
                        await websocket.send_json({
                            "type": "terminate_result",
                            "call_id": call_id,
                            "success": True
                        })
                    else:
                        await websocket.send_json({
                            "type": "terminate_result",
                            "call_id": call_id,
                            "success": False,
                            "error": "Llamada no encontrada"
                        })
            
            except json.JSONDecodeError:
                print("Error al decodificar mensaje JSON")
            except Exception as e:
                print(f"Error procesando mensaje: {str(e)}")
    
    except WebSocketDisconnect:
        ws_manager.disconnect(websocket)
        print(f"Cliente WebSocket desconectado. Conexiones restantes: {len(ws_manager.active_connections)}")
    except Exception as e:
        print(f"Error en WebSocket: {str(e)}")
        ws_manager.disconnect(websocket)    

@app.post("/api/active-calls")
async def report_active_call(call_data: dict):
    print(f"Recibido reporte de llamada activa: {call_data}")
    
    try:
        # Extraer datos
        call_id = call_data.get("call_id")
        
        if not call_id:
            return {"status": "error", "message": "call_id es requerido"}
        
        # Mapear campos según sea necesario
        db_call = {
            "call_id": call_id,
            "calling_number": call_data.get("calling_number") or call_data.get("origin"),
            "called_number": call_data.get("called_number") or call_data.get("destination"),
            "start_time": datetime.fromisoformat(call_data.get("start_time").replace('Z', '+00:00')) 
                if call_data.get("start_time") else datetime.now(),
            "last_updated": datetime.now(),
            "current_duration": call_data.get("current_duration") or call_data.get("duration", 0),
            "current_cost": call_data.get("current_cost") or call_data.get("estimatedCost", 0),
            "connection_id": call_data.get("connection_id") or call_id
        }
        
        # Actualizar en la base de datos
        db = SessionLocal()
        
        try:
            # Verificar si ya existe
            existing = db.execute(
                text("SELECT id FROM active_calls WHERE call_id = :call_id"),
                {"call_id": call_id}
            ).fetchone()
            
            if existing:
                # Actualizar
                set_clause = ", ".join([f"{k} = :{k}" for k in db_call.keys()])
                query = text(f"UPDATE active_calls SET {set_clause} WHERE call_id = :call_id")
                db.execute(query, db_call)
            else:
                # Insertar nuevo
                columns = ", ".join(db_call.keys())
                placeholders = ", ".join([f":{k}" for k in db_call.keys()])
                query = text(f"INSERT INTO active_calls ({columns}) VALUES ({placeholders})")
                db.execute(query, db_call)
            
            db.commit()
            
            db = SessionLocal()
            active_calls_rows = db.execute(
                text("SELECT * FROM active_calls")
            ).fetchall()
            
            # Convertir a formato que el cliente espera
            active_calls = []
            for row in active_calls_rows:
                # Asumiendo que las columnas están en este orden
                call = {
                    "call_id": row[1],  # call_id
                    "calling_number": row[2],  # calling_number
                    "called_number": row[3],  # called_number
                    "start_time": row[4].isoformat() if row[4] else None,  # start_time
                    "current_duration": row[6] if row[6] is not None else 0,  # current_duration
                    "current_cost": float(row[7]) if row[7] is not None else 0.0  # current_cost
                }
                active_calls.append(call)
            
            # Broadcast con debug
            if ws_manager.active_connections:
                message = {
                    "type": "update",
                    "active_calls": active_calls
                }
                print(f"Enviando broadcast con {len(active_calls)} llamadas activas: {active_calls}")
                await ws_manager.broadcast(message)
                print(f"Broadcast enviado a {len(ws_manager.active_connections)} conexiones")
            
            return {"status": "ok", "active_calls_count": len(active_calls)}
                
        except Exception as e:
            db.rollback()
            print(f"Error en BD: {str(e)}")
            raise
        finally:
            db.close()
            
    except Exception as e:
        print(f"Error general: {str(e)}")
        return {"status": "error", "message": str(e)}

@app.delete("/api/active-calls/{call_id}")
async def remove_active_call(call_id: str):
    try:
        db = SessionLocal()
        
        # Buscar primero por call_id
        result = db.execute(
            text("SELECT id FROM active_calls WHERE call_id = :call_id"),
            {"call_id": call_id}
        ).fetchone()
        
        # Si no se encuentra, buscar por connection_id como alternativa
        if not result:
            result = db.execute(
                text("SELECT id FROM active_calls WHERE connection_id = :call_id"),
                {"call_id": call_id}
            ).fetchone()
        
        if result:
            # Eliminar la llamada
            db.execute(
                text("DELETE FROM active_calls WHERE id = :id"),
                {"id": result[0]}
            )
            db.commit()
            print(f"Llamada eliminada: {call_id}")
            
            # Obtener lista actualizada para broadcast
            active_calls_rows = db.execute(
                text("SELECT call_id, calling_number, called_number, start_time, current_duration, current_cost FROM active_calls")
            ).fetchall()
            
            # Broadcast a clientes WebSocket
            if ws_manager.active_connections:
                active_calls = []
                for row in active_calls_rows:
                    call = {
                        "call_id": row[0],
                        "calling_number": row[1],
                        "called_number": row[2],
                        "start_time": row[3].isoformat() if row[3] else None,
                        "current_duration": row[4] if row[4] is not None else 0,
                        "current_cost": float(row[5]) if row[5] is not None else 0.0
                    }
                    active_calls.append(call)
                
                await ws_manager.broadcast({
                    "type": "update",
                    "active_calls": active_calls
                })
            
            return {"status": "ok", "message": f"Llamada {call_id} eliminada correctamente"}
        else:
            print(f"No se encontró la llamada a eliminar: {call_id}")
            return {"status": "not_found", "message": f"Llamada {call_id} no encontrada"}
        
    except Exception as e:
        print(f"Error al eliminar llamada activa: {str(e)}")
        return {"status": "error", "message": str(e)}
        
# Endpoints para estadísticas y monitoreo
@app.get("/api/ws-stats")
async def get_ws_stats():
    return {
        **ws_stats,
        "active_connections": ws_manager.connection_count,
        "active_calls_count": len(active_calls),
        "timestamp": datetime.now().isoformat()
    }

class CDR(Base):
    __tablename__ = "cdr"
    id = Column(Integer, primary_key=True, index=True)
    calling_number = Column(String)
    called_number = Column(String)
    start_time = Column(DateTime)
    end_time = Column(DateTime)
    duration_seconds = Column(Integer)
    duration_billable = Column(Integer)  # Nuevo campo
    cost = Column(Numeric(10,2))
    status = Column(String)
    direction = Column(String)  # Ya existe pero no se está mapeando
    release_cause = Column(Integer)
    connect_time = Column(DateTime)  # Ya existe pero se llama answer_time en Java
    dialing_time = Column(DateTime)  # Nuevo campo
    network_reached_time = Column(DateTime)  # Nuevo campo
    network_alerting_time = Column(DateTime)  # Nuevo campo
    zona_id = Column(Integer)  # Ya existe

class SaldoAnexo(Base):
    __tablename__ = "saldo_anexos"
    id = Column(Integer, primary_key=True, index=True)
    calling_number = Column(String, unique=True)
    saldo = Column(Numeric(10,2))
    fecha_ultima_recarga = Column(DateTime, default=datetime.utcnow)

class Recarga(Base):
    __tablename__ = "recargas"
    id = Column(Integer, primary_key=True, index=True)
    calling_number = Column(String)
    monto = Column(Numeric(10,2))
    fecha = Column(DateTime, default=datetime.utcnow)

class Auditoria(Base):
    __tablename__ = "saldo_auditoria"
    id = Column(Integer, primary_key=True, index=True)
    calling_number = Column(String)
    saldo_anterior = Column(Numeric(10,2))
    saldo_nuevo = Column(Numeric(10,2))
    tipo_accion = Column(String)
    fecha = Column(DateTime, default=datetime.utcnow)

class Anexo(Base):
    __tablename__ = "anexos"
    id = Column(Integer, primary_key=True, index=True)
    numero = Column(String, unique=True, index=True)
    usuario = Column(String)
    area_nivel1 = Column(String)
    area_nivel2 = Column(String, nullable=True)
    area_nivel3 = Column(String, nullable=True)
    pin = Column(String)  # Almacenaremos un hash del PIN
    saldo_actual = Column(Numeric(10, 2), default=0)
    fecha_creacion = Column(DateTime, default=datetime.utcnow)
    activo = Column(Boolean, default=True)

class Configuracion(Base):
    __tablename__ = "configuracion"
    id = Column(Integer, primary_key=True, index=True)
    clave = Column(String, unique=True)
    valor = Column(String)
    descripcion = Column(String, nullable=True)

# Agregar al principio de main.py, junto con las otras definiciones de modelos
class CucmConfig(Base):
    __tablename__ = "cucm_config"
    id = Column(Integer, primary_key=True, index=True)
    server_ip = Column(String, nullable=False)
    server_port = Column(Integer, default=2748)
    username = Column(String, nullable=False)
    password = Column(String, nullable=False)
    app_info = Column(String, default="TarificadorApp")
    reconnect_delay = Column(Integer, default=30)
    check_interval = Column(Integer, default=60)
    enabled = Column(Boolean, default=True)
    last_updated = Column(DateTime, default=datetime.utcnow)
    last_status = Column(String, default="unknown")
    last_status_update = Column(DateTime, nullable=True)

class Usuario(Base):
    __tablename__ = "usuarios"
    id = Column(Integer, primary_key=True, index=True)
    username = Column(String, unique=True, index=True)
    password = Column(String)
    nombre = Column(String, nullable=True)
    apellido = Column(String, nullable=True)
    email = Column(String, nullable=True)
    role = Column(String)
    activo = Column(Boolean, default=True)
    ultimo_login = Column(DateTime, nullable=True)

# Nuevos modelos para gestión de zonas y tarifas
class Zona(Base):
    __tablename__ = "zonas"
    id = Column(Integer, primary_key=True, index=True)
    nombre = Column(String, unique=True)
    descripcion = Column(String)
    prefijos = relationship("Prefijo", back_populates="zona")
    tarifas = relationship("Tarifa", back_populates="zona")

class Prefijo(Base):
    __tablename__ = "prefijos"
    id = Column(Integer, primary_key=True, index=True)
    zona_id = Column(Integer, ForeignKey("zonas.id"))
    prefijo = Column(String)
    longitud_minima = Column(Integer)
    longitud_maxima = Column(Integer)
    zona = relationship("Zona", back_populates="prefijos")

class Tarifa(Base):
    __tablename__ = "tarifas"
    id = Column(Integer, primary_key=True, index=True)
    zona_id = Column(Integer, ForeignKey("zonas.id"))
    tarifa_segundo = Column(Numeric(10, 5))  # 5 decimales de precisión
    fecha_inicio = Column(DateTime, default=datetime.utcnow)
    activa = Column(Boolean, default=True)
    zona = relationship("Zona", back_populates="tarifas")

Base.metadata.create_all(bind=engine)

# Función para inicializar zonas y prefijos
def inicializar_zonas_y_prefijos():
    db = SessionLocal()
    
    # Verificar si ya existen zonas
    check_query = text("SELECT COUNT(*) FROM zonas")
    count = db.execute(check_query).scalar()
    
    if count > 0:
        db.close()
        return
    
    # Crear zonas iniciales
    zonas = [
        {"nombre": "Local", "descripcion": "Llamadas locales de 7 dígitos", "tarifa": 0.00015},
        {"nombre": "Movil", "descripcion": "Llamadas a celulares", "tarifa": 0.00030},
        {"nombre": "LDN", "descripcion": "Larga Distancia Nacional", "tarifa": 0.00050},
        {"nombre": "LDI", "descripcion": "Larga Distancia Internacional", "tarifa": 0.00120},
        {"nombre": "Emergencia", "descripcion": "Números de emergencia", "tarifa": 0.00000},
        {"nombre": "0800", "descripcion": "Números gratuitos", "tarifa": 0.00000}
    ]
    
    zonas_ids = {}
    
    for zona_data in zonas:
        # Insertar zona
        insert_zona_query = text("""
            INSERT INTO zonas (nombre, descripcion) 
            VALUES (:nombre, :descripcion)
            RETURNING id
        """)
        
        result = db.execute(insert_zona_query, {
            "nombre": zona_data["nombre"],
            "descripcion": zona_data["descripcion"]
        })
        
        zona_id = result.fetchone()[0]
        zonas_ids[zona_data["nombre"]] = zona_id
        
        # Insertar tarifa para la zona
        insert_tarifa_query = text("""
            INSERT INTO tarifas (zona_id, tarifa_segundo, fecha_inicio, activa)
            VALUES (:zona_id, :tarifa_segundo, CURRENT_TIMESTAMP, TRUE)
        """)
        
        db.execute(insert_tarifa_query, {
            "zona_id": zona_id,
            "tarifa_segundo": zona_data["tarifa"]
        })
    
    # Insertar prefijos
    prefijos = [
        # Local - números fijos 7 dígitos (2-9)XXXXXX
        {"zona_id": zonas_ids["Local"], "prefijo": "2", "longitud_minima": 7, "longitud_maxima": 7},
        {"zona_id": zonas_ids["Local"], "prefijo": "3", "longitud_minima": 7, "longitud_maxima": 7},
        {"zona_id": zonas_ids["Local"], "prefijo": "4", "longitud_minima": 7, "longitud_maxima": 7},
        {"zona_id": zonas_ids["Local"], "prefijo": "5", "longitud_minima": 7, "longitud_maxima": 7},
        {"zona_id": zonas_ids["Local"], "prefijo": "6", "longitud_minima": 7, "longitud_maxima": 7},
        {"zona_id": zonas_ids["Local"], "prefijo": "7", "longitud_minima": 7, "longitud_maxima": 7},
        {"zona_id": zonas_ids["Local"], "prefijo": "8", "longitud_minima": 7, "longitud_maxima": 7},
        {"zona_id": zonas_ids["Local"], "prefijo": "9", "longitud_minima": 7, "longitud_maxima": 7},
        
        # Móvil - 9 dígitos 9XXXXXXXX
        {"zona_id": zonas_ids["Movil"], "prefijo": "9", "longitud_minima": 9, "longitud_maxima": 9},
        
        # LDN - 0[4-8]XXXXXXX
        {"zona_id": zonas_ids["LDN"], "prefijo": "04", "longitud_minima": 9, "longitud_maxima": 10},
        {"zona_id": zonas_ids["LDN"], "prefijo": "05", "longitud_minima": 9, "longitud_maxima": 10},
        {"zona_id": zonas_ids["LDN"], "prefijo": "06", "longitud_minima": 9, "longitud_maxima": 10},
        {"zona_id": zonas_ids["LDN"], "prefijo": "07", "longitud_minima": 9, "longitud_maxima": 10},
        {"zona_id": zonas_ids["LDN"], "prefijo": "08", "longitud_minima": 9, "longitud_maxima": 10},
        
        # LDI - 00[1-9]XXXXXXX.... (10-15 dígitos)
        {"zona_id": zonas_ids["LDI"], "prefijo": "001", "longitud_minima": 10, "longitud_maxima": 15},
        {"zona_id": zonas_ids["LDI"], "prefijo": "002", "longitud_minima": 10, "longitud_maxima": 15},
        {"zona_id": zonas_ids["LDI"], "prefijo": "003", "longitud_minima": 10, "longitud_maxima": 15},
        {"zona_id": zonas_ids["LDI"], "prefijo": "004", "longitud_minima": 10, "longitud_maxima": 15},
        {"zona_id": zonas_ids["LDI"], "prefijo": "005", "longitud_minima": 10, "longitud_maxima": 15},
        {"zona_id": zonas_ids["LDI"], "prefijo": "006", "longitud_minima": 10, "longitud_maxima": 15},
        {"zona_id": zonas_ids["LDI"], "prefijo": "007", "longitud_minima": 10, "longitud_maxima": 15},
        {"zona_id": zonas_ids["LDI"], "prefijo": "008", "longitud_minima": 10, "longitud_maxima": 15},
        {"zona_id": zonas_ids["LDI"], "prefijo": "009", "longitud_minima": 10, "longitud_maxima": 15},
        
        # Emergencia 1XX
        {"zona_id": zonas_ids["Emergencia"], "prefijo": "1", "longitud_minima": 3, "longitud_maxima": 3},
        
        # 0800 - 0800XXXXX
        {"zona_id": zonas_ids["0800"], "prefijo": "0800", "longitud_minima": 9, "longitud_maxima": 9},
    ]
    
    for prefijo_data in prefijos:
        insert_prefijo_query = text("""
            INSERT INTO prefijos (zona_id, prefijo, longitud_minima, longitud_maxima)
            VALUES (:zona_id, :prefijo, :longitud_minima, :longitud_maxima)
        """)
        
        db.execute(insert_prefijo_query, {
            "zona_id": prefijo_data["zona_id"],
            "prefijo": prefijo_data["prefijo"],
            "longitud_minima": prefijo_data["longitud_minima"],
            "longitud_maxima": prefijo_data["longitud_maxima"]
        })
    
    db.commit()
    db.close()

def determinar_zona_y_tarifa(numero_marcado: str, db):
    """
    Determina la zona del número marcado y obtiene la tarifa correspondiente.
    Usa SQLAlchemy ORM con los modelos Zona, Prefijo y Tarifa.
    """
    # Limpiar el número (quitar caracteres especiales)
    numero_limpio = ''.join(filter(str.isdigit, numero_marcado))
    longitud_numero = len(numero_limpio)
    
    # Buscar prefijos que coincidan con la longitud del número
    prefijos_candidatos = db.query(Prefijo).filter(
        Prefijo.longitud_minima <= longitud_numero,
        Prefijo.longitud_maxima >= longitud_numero
    ).all()
    
    # Encontrar el prefijo más específico (más largo) que coincida
    mejor_prefijo = None
    mejor_longitud = 0
    
    for prefijo_obj in prefijos_candidatos:
        # Verificar si el número comienza con este prefijo
        if numero_limpio.startswith(prefijo_obj.prefijo):
            # Si este prefijo es más específico (más largo), usarlo
            if len(prefijo_obj.prefijo) > mejor_longitud:
                mejor_prefijo = prefijo_obj
                mejor_longitud = len(prefijo_obj.prefijo)
    
    if mejor_prefijo:
        # Obtener la tarifa activa para esta zona usando la relación
        tarifa_activa = db.query(Tarifa).filter(
            Tarifa.zona_id == mejor_prefijo.zona_id,
            Tarifa.activa == True
        ).order_by(Tarifa.fecha_inicio.desc()).first()
        
        if tarifa_activa:
            return {
                'prefijo_id': mejor_prefijo.id,
                'zona_id': mejor_prefijo.zona_id,
                'prefijo': mejor_prefijo.prefijo,
                'zona_nombre': mejor_prefijo.zona.nombre,
                'zona_descripcion': mejor_prefijo.zona.descripcion,
                'tarifa_segundo': float(tarifa_activa.tarifa_segundo),
                'tarifa_id': tarifa_activa.id,
                'numero_valido': True
            }
        else:
            # No hay tarifa activa para esta zona
            return {
                'prefijo_id': mejor_prefijo.id,
                'zona_id': mejor_prefijo.zona_id,
                'prefijo': mejor_prefijo.prefijo,
                'zona_nombre': mejor_prefijo.zona.nombre,
                'zona_descripcion': f"Sin tarifa activa: {mejor_prefijo.zona.descripcion}",
                'tarifa_segundo': 0.0,
                'tarifa_id': None,
                'numero_valido': False
            }
    
    # Si no se encuentra ningún prefijo que coincida
    return {
        'prefijo_id': None,
        'zona_id': None,
        'prefijo': 'UNKNOWN',
        'zona_nombre': 'Desconocida',
        'zona_descripcion': f'Número no reconocido: {numero_marcado} (longitud: {longitud_numero})',
        'tarifa_segundo': 0.0,
        'tarifa_id': None,
        'numero_valido': False
    }

@app.get("/check_balance_for_call/{calling_number}/{called_number}")
def check_balance_for_call(calling_number: str, called_number: str):
    """
    Verifica si hay saldo suficiente para realizar una llamada específica.
    Sistema basado completamente en segundos.
    """
    try:
        # Obtener saldo actual
        with SessionLocal() as db:
            saldo_anexo = db.query(SaldoAnexo).filter(
                SaldoAnexo.calling_number == calling_number
            ).first()
            
            if not saldo_anexo:
                return {
                    "has_balance": False, 
                    "balance": 0.0, 
                    "can_call": False, 
                    "reason": "No account found"
                }
            
            saldo_actual = float(saldo_anexo.saldo)
            
            # Determinar zona y tarifa para el número de destino
            zona_info = determinar_zona_y_tarifa(called_number, db)
            
            if not zona_info['numero_valido']:
                return {
                    "has_balance": saldo_actual > 0,
                    "balance": saldo_actual,
                    "can_call": False,
                    "reason": f"Invalid destination number: {called_number}"
                }
            
            # Todo basado en segundos
            tarifa_segundo = zona_info['tarifa_segundo']
            
            # Verificar si puede realizar al menos 1 segundo de llamada
            can_call = saldo_actual >= tarifa_segundo
            
            # Calcular tiempo disponible en segundos
            tiempo_disponible_segundos = int(saldo_actual / tarifa_segundo) if tarifa_segundo > 0 else 999999
            
            return {
                "has_balance": saldo_actual > 0,
                "balance": saldo_actual,
                "can_call": can_call,
                "zona": zona_info['zona_nombre'],
                "tarifa_segundo": tarifa_segundo,
                "tiempo_disponible_segundos": tiempo_disponible_segundos
            }
        
    except Exception as e:
        print(f"Error verificando saldo para llamada: {e}")
        import traceback
        traceback.print_exc()
        return {
            "has_balance": False, 
            "balance": 0.0, 
            "can_call": False, 
            "reason": str(e)
        }
    
# Llamar a la función de inicialización al arrancar la aplicación
@app.on_event("startup")
def startup_event():
    inicializar_zonas_y_prefijos()

# Función para determinar la zona de un número
def determinar_zona(numero):
    db = SessionLocal()
    
    # Consultar todos los prefijos ordenados por longitud descendente
    query = text("""
        SELECT id, zona_id, prefijo, longitud_minima, longitud_maxima 
        FROM prefijos 
        ORDER BY LENGTH(prefijo) DESC
    """)
    
    prefijos = db.execute(query).fetchall()
    db.close()
    
    for prefijo in prefijos:
        prefijo_str = prefijo[2]
        longitud_minima = prefijo[3]
        longitud_maxima = prefijo[4]
        
        if (numero.startswith(prefijo_str) and 
            longitud_minima <= len(numero) <= longitud_maxima):
            return prefijo[1]  # Retornar zona_id
    
    return None  # Si no se encuentra una zona

# Función para obtener la tarifa activa de una zona
def obtener_tarifa(zona_id):
    if not zona_id:
        return 0.0005  # Tarifa por defecto si no se encuentra zona
    
    db = SessionLocal()
    
    query = text("""
        SELECT tarifa_segundo 
        FROM tarifas 
        WHERE zona_id = :zona_id AND activa = TRUE 
        ORDER BY fecha_inicio DESC 
        LIMIT 1
    """)
    
    result = db.execute(query, {"zona_id": zona_id}).fetchone()
    db.close()
    
    if result:
        return float(result[0])
    
    return 0.0005  # Tarifa por defecto si no hay tarifa activa

@manager.user_loader()
def load_user(username: str):
    try:
        # Usar SQLAlchemy para consultar el usuario en la tabla 'usuarios'
        db = SessionLocal()
        # Usamos text para ejecutar una consulta SQL directa
        query = text("SELECT username, password, role FROM usuarios WHERE username = :username")
        result = db.execute(query, {"username": username}).fetchone()
        
        if result:
            # Devolver un diccionario con los datos del usuario
            user_dict = {
                "username": result[0],
                "password": result[1],
                "role": result[2]
            }
            db.close()
            return user_dict
        else:
            db.close()
            return None
    except Exception as e:
        print(f"Error cargando usuario desde DB: {e}")
        return None
        
async def authenticated_user(request: Request):
    token = request.cookies.get(manager.cookie_name)
    if token is None:
        return RedirectResponse(url="/login")
    try:
        user = await manager.get_current_user(token)
    except Exception:
        return RedirectResponse(url="/login")
    return user

async def admin_only(request: Request):
    user = await authenticated_user(request)
    # Si authenticated_user devuelve un RedirectResponse, simplemente devuélvelo
    if isinstance(user, RedirectResponse):
        return user
    if user["role"] != "admin":
        return RedirectResponse(url="/login")
    return user

@app.post("/auth/login")
def login(request: Request, username: str = Form(...), password: str = Form(...)):
    user = load_user(username)
    if not user or not pwd_context.verify(password, user['password']):
        return templates.TemplateResponse("login.html", {"request": request, "error": "Credenciales inválidas"})

    # Actualizar último login
    try:
        db = SessionLocal()
        update_query = text("UPDATE usuarios SET ultimo_login = CURRENT_TIMESTAMP WHERE username = :username")
        db.execute(update_query, {"username": username})
        db.commit()
        db.close()
    except Exception as e:
        print(f"Error actualizando último login: {e}")

    access_token = manager.create_access_token(data={"sub": username})
    resp = RedirectResponse(url="/dashboard/saldo", status_code=302)
    manager.set_cookie(resp, access_token)
    return resp

@app.get("/login")
def login_page(request: Request):
    return templates.TemplateResponse("login.html", {"request": request})

# Modelos para entrada de datos
from pydantic import BaseModel
from fastapi import BackgroundTasks
from fastapi.staticfiles import StaticFiles
from weasyprint import HTML
from jinja2 import Template

class CallEvent(BaseModel):
    calling_number: str
    called_number: str
    start_time: datetime
    end_time: datetime
    duration_seconds: int
    duration_billable: int
    status: str
    direction: str
    release_cause: int
    answer_time: datetime = None  # Campo opcional
    dialing_time: datetime = None
    network_reached_time: datetime = None
    network_alerting_time: datetime = None

# Nuevos modelos para zonas y tarifas
class ZonaCreate(BaseModel):
    nombre: str
    descripcion: str

class PrefijoCreate(BaseModel):
    zona_id: int
    prefijo: str
    longitud_minima: int
    longitud_maxima: int

class TarifaCreate(BaseModel):
    zona_id: int
    tarifa_segundo: float

# API Principal - Modificada para usar zonas y tarifas
@app.post("/cdr")
def create_cdr(event: CallEvent):
    db = SessionLocal()
    
    # Calcular el costo basado en duration_billable (duración facturable)
    cost = (event.duration_billable / 60) * 0.05  # 5 centavos por minuto
    
    # Crear el CDR mapeando answer_time a connect_time
    cdr_data = {
        "calling_number": event.calling_number,
        "called_number": event.called_number,
        "start_time": event.start_time,
        "end_time": event.end_time,
        "duration_seconds": event.duration_seconds,
        "duration_billable": event.duration_billable,
        "cost": cost,
        "status": event.status,
        "direction": event.direction,
        "release_cause": event.release_cause,
        "connect_time": event.answer_time,  # Mapear answer_time a connect_time
        "dialing_time": event.dialing_time,
        "network_reached_time": event.network_reached_time,
        "network_alerting_time": event.network_alerting_time,
        "zona_id": 2  # O calcular según la lógica de tu sistema
    }
    
    cdr = CDR(**cdr_data)
    db.add(cdr)
    
    # Actualizar el saldo (usando el costo calculado)
    db.execute(
        text("UPDATE saldo_anexos SET saldo = saldo - :cost WHERE calling_number = :calling_number"),
        {"cost": cost, "calling_number": event.calling_number}
    )
    
    # Verificar saldo bajo
    nuevo_saldo = db.execute(
        text("SELECT saldo FROM saldo_anexos WHERE calling_number = :calling_number"),
        {"calling_number": event.calling_number}
    ).fetchone()
    
    if nuevo_saldo and nuevo_saldo[0] < 1.0:
        # Aquí puedes programar enviar correo o alerta
        pass
    
    db.commit()
    db.close()
    return {"message": "CDR saved", "cost": cost}


# Nuevo endpoint para verificar saldo con destino
@app.get("/check_balance/{calling_number}/{called_number}")
def check_balance_with_destination(calling_number: str, called_number: str):
    db = SessionLocal()
    
    # Verificar saldo
    query = text("SELECT saldo FROM saldo_anexos WHERE calling_number = :calling_number")
    saldo = db.execute(query, {"calling_number": calling_number}).fetchone()
    
    if not saldo or saldo[0] <= 0:
        db.close()
        return {"has_balance": False, "reason": "insufficient_balance"}
    
    # Determinar la zona y tarifa del destino
    zona_id = determinar_zona(called_number)
    if not zona_id:
        db.close()
        return {"has_balance": False, "reason": "invalid_destination"}
    
    # Verificar si el saldo es suficiente para al menos 1 minuto de llamada
    tarifa_segundo = obtener_tarifa(zona_id)
    costo_minuto = tarifa_segundo * 60
    
    if saldo[0] < costo_minuto:
        db.close()
        return {"has_balance": False, "reason": "low_balance_for_call"}
    
    db.close()
    return {"has_balance": True, "zona_id": zona_id, "tarifa_segundo": tarifa_segundo}

# Mantener compatibilidad con la versión anterior
@app.get("/check_balance/{calling_number}")
def check_balance(calling_number: str):
    with SessionLocal() as db:
        query = text("SELECT saldo FROM saldo_anexos WHERE calling_number = :calling_number")
        saldo = db.execute(query, {"calling_number": calling_number}).fetchone()
        return {"has_balance": True}
        #return {"has_balance": saldo and saldo[0] > 0}

@app.post("/recargar/{calling_number}/{amount}")
async def recargar_saldo(calling_number: str, amount: float, user=Depends(admin_only)):
    db = SessionLocal()
    
    # Check if the record exists
    query = text("SELECT saldo FROM saldo_anexos WHERE calling_number = :calling_number")
    saldo_actual = db.execute(query, {"calling_number": calling_number}).fetchone()

    if saldo_actual:
        update_query = text(
            "UPDATE saldo_anexos SET saldo = saldo + :amount, fecha_ultima_recarga = CURRENT_TIMESTAMP WHERE calling_number = :calling_number"
        )
        db.execute(update_query, {"amount": amount, "calling_number": calling_number})
    else:
        insert_query = text(
            "INSERT INTO saldo_anexos (calling_number, saldo, fecha_ultima_recarga) VALUES (:calling_number, :amount, CURRENT_TIMESTAMP)"
        )
        db.execute(insert_query, {"calling_number": calling_number, "amount": amount})

    # Logging the audit record  
    audit_query = text(
        "INSERT INTO saldo_auditoria (calling_number, saldo_anterior, saldo_nuevo, tipo_accion) "
        "VALUES (:calling_number, :saldo_anterior, :saldo_nuevo, 'recarga')"
    )
    db.execute(audit_query, {
        "calling_number": calling_number, 
        "saldo_anterior": saldo_actual[0] if saldo_actual else 0,
        "saldo_nuevo": saldo_actual[0] + Decimal(str(amount)) if saldo_actual else Decimal(str(amount))
        #"saldo_nuevo": saldo_actual[0] + amount if saldo_actual else amount
    })

    # Record the recharge
    recharge_query = text("INSERT INTO recargas (calling_number, monto) VALUES (:calling_number, :amount)")
    db.execute(recharge_query, {"calling_number": calling_number, "amount": amount})

    db.commit()
    db.close()
    return {"message": f"Recargado {amount} al número {calling_number}"}

# DASHBOARD: SALDO
@app.get("/dashboard/saldo")
async def dashboard_saldo(request: Request, user=Depends(authenticated_user)):
    db = SessionLocal()
    query = text("SELECT calling_number, saldo FROM saldo_anexos ORDER BY calling_number ASC")
    rows = db.execute(query).fetchall()
    saldos_bajos = [row for row in rows if float(row[1]) < 1.0]
    labels = [row[0] for row in rows]
    data = [float(row[1]) for row in rows]
    
    # Pre-calculate the background colors based on saldo values
    background_colors = ['rgba(220, 53, 69, 0.8)' if float(row[1]) < 1.0 else 'rgba(13, 110, 253, 0.8)' for row in rows]
    
    db.close()
    return templates.TemplateResponse("dashboard_saldo.html", {
        "request": request, 
        "rows": rows, 
        "labels": labels, 
        "data": data, 
        "saldos_bajos": saldos_bajos, 
        "user": user, 
        "background_colors": background_colors
    })

# DASHBOARD: CDR - Modificado para incluir zona
@app.get("/dashboard/cdr")
def dashboard_cdr(request: Request, 
                  user=Depends(authenticated_user),
                  page: int = Query(1, ge=1),
                  per_page: int = Query(10, ge=1, le=100),
                  calling_number: str = Query(None),
                  start_date: str = Query(None),
                  end_date: str = Query(None),
                  min_duration: int = Query(0)):

    db = SessionLocal()
    offset = (page - 1) * per_page
    
    # Consulta base incluyendo duration_billable
    query = "SELECT calling_number, called_number, start_time, end_time, duration_seconds, cost, duration_billable FROM cdr WHERE 1=1"
    
    # Parámetros para la consulta parametrizada
    params = {}
    
    # Aplicar filtros
    if calling_number:
        query += " AND calling_number = :calling_number"
        params['calling_number'] = calling_number
    
    if start_date:
        query += " AND start_time >= :start_date"
        params['start_date'] = f"{start_date} 00:00:00"
    
    if end_date:
        query += " AND start_time <= :end_date"
        params['end_date'] = f"{end_date} 23:59:59"
    
    if min_duration:
        query += " AND duration_seconds >= :min_duration"
        params['min_duration'] = min_duration
    
    # Ordenar y paginar
    query += " ORDER BY start_time DESC LIMIT :limit OFFSET :offset"
    params['limit'] = per_page
    params['offset'] = offset
    
    # Ejecutar consulta
    rows = db.execute(text(query), params).fetchall()
    
    # No necesitamos procesar las filas ya que duration_billable viene de la BD
    processed_rows = [tuple(row) for row in rows]
    
    # Calcular estadísticas para el gráfico de barras
    horas = [row[2].hour for row in rows]
    contador = Counter(horas)
    labels = sorted(contador.keys())
    data = [contador[h] for h in labels]
    
    # Calcular total de páginas
    count_query = "SELECT COUNT(*) FROM cdr WHERE 1=1"
    count_params = {}
    
    if calling_number:
        count_query += " AND calling_number = :calling_number"
        count_params['calling_number'] = calling_number
    
    if start_date:
        count_query += " AND start_time >= :start_date"
        count_params['start_date'] = f"{start_date} 00:00:00"
    
    if end_date:
        count_query += " AND start_time <= :end_date"
        count_params['end_date'] = f"{end_date} 23:59:59"
    
    if min_duration:
        count_query += " AND duration_seconds >= :min_duration"
        count_params['min_duration'] = min_duration
    
    total_records = db.execute(text(count_query), count_params).fetchone()[0]
    total_pages = -(-total_records // per_page)  # División entera redondeada hacia arriba
    
    db.close()
    
    return templates.TemplateResponse("dashboard_cdr.html", {
        "request": request, 
        "rows": processed_rows,
        "labels": labels, 
        "data": data, 
        "page": page, 
        "per_page": per_page, 
        "total_pages": total_pages,
        "user": user
    })

@app.get("/dashboard/finanzas")
async def dashboard_finanzas(request: Request, user=Depends(authenticated_user)):
    db = SessionLocal()
    query = text("""
        SELECT DATE(start_time) as fecha, SUM(cost) as total_cost
        FROM cdr
        GROUP BY DATE(start_time)
        ORDER BY fecha DESC
        LIMIT 30
    """)
    rows = db.execute(query).fetchall()
    db.close()

    labels = [str(row[0]) for row in rows]
    data = [float(row[1]) for row in rows]

    return templates.TemplateResponse("dashboard_finanzas.html", {
        "request": request, "labels": labels, "data": data, "user": user
    })


@app.get("/dashboard/recargas")
async def dashboard_recargas(request: Request, user=Depends(authenticated_user)):
    db = SessionLocal()
    query = text("SELECT calling_number, monto, fecha FROM recargas ORDER BY fecha DESC LIMIT 100")
    rows = db.execute(query).fetchall()
    db.close()

    return templates.TemplateResponse("dashboard_recargas.html", {
        "request": request, "rows": rows, "user": user
    })

@app.get("/dashboard/auditoria")
async def dashboard_auditoria(request: Request, user=Depends(admin_only)):
    db = SessionLocal()
    query = text("SELECT calling_number, saldo_anterior, saldo_nuevo, tipo_accion, fecha FROM saldo_auditoria ORDER BY fecha DESC LIMIT 100")
    rows = db.execute(query).fetchall()
    db.close()

    return templates.TemplateResponse("dashboard_auditoria.html", {
        "request": request, "rows": rows, "user": user
    })

@app.get("/dashboard/ranking_consumo")
async def dashboard_ranking_consumo(request: Request, user=Depends(authenticated_user)):
    db = SessionLocal()
    query = text("""
        SELECT calling_number, SUM(cost) as consumo_total
        FROM cdr
        GROUP BY calling_number
        ORDER BY consumo_total DESC
        LIMIT 10
    """)
    rows = db.execute(query).fetchall()
    db.close()

    labels = [row[0] for row in rows]
    data = [float(row[1]) for row in rows]
    
    # Calcular la suma total para evitar división por cero en la plantilla
    total_consumo = sum(data) if data else 0
    
    # Calcular porcentajes de forma segura
    porcentajes = []
    if total_consumo > 0 and rows:
        porcentajes = [(float(row[1]) / total_consumo * 100) for row in rows]
    
    return templates.TemplateResponse("dashboard_ranking_consumo.html", {
        "request": request, 
        "labels": labels, 
        "data": data, 
        "rows": rows, 
        "user": user,
        "total_consumo": total_consumo,
        "porcentajes": porcentajes
    })

@app.get("/export/saldo/pdf")
async def export_saldo_pdf(user=Depends(admin_only)):
    db = SessionLocal()
    query = text("SELECT calling_number, saldo FROM saldo_anexos ORDER BY calling_number ASC")
    rows = db.execute(query).fetchall()
    db.close()

    html_template = """
    <html>
    <body>
    <h1>Reporte de Saldos</h1>
    <table border="1">
        <thead>
            <tr><th>Anexo</th><th>Saldo</th></tr>
        </thead>
        <tbody>
        {% for row in rows %}
            <tr><td>{{ row[0] }}</td><td>${{ "%.2f"|format(row[1]) }}</td></tr>
        {% endfor %}
        </tbody>
    </table>
    </body>
    </html>
    """
    template = Template(html_template)
    html_content = template.render(rows=rows)

    pdf = HTML(string=html_content).write_pdf()

    return StreamingResponse(
        io.BytesIO(pdf),
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=saldo_report.pdf"}
    )


@app.get("/dashboard/recarga_masiva")
async def form_recarga_masiva(request: Request, user=Depends(admin_only)):
    if isinstance(user, RedirectResponse):
        return user
    return templates.TemplateResponse("recarga_masiva.html", {"request": request, "user": user})

@app.post("/dashboard/recarga_masiva")
async def recarga_masiva(request: Request, file: UploadFile = File(...), user=Depends(admin_only)):
    db = SessionLocal()
    
    try:
        # Leer el contenido del archivo
        content = await file.read()
        
        # Determinar tipo de archivo por extensión
        filename = file.filename.lower()
        
        # Lista para almacenar las filas a procesar
        rows = []
        
        # Procesar según el tipo de archivo
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            # Archivo Excel
            try:
                import io
                import openpyxl
                
                # Cargar el archivo Excel
                wb = openpyxl.load_workbook(io.BytesIO(content))
                ws = wb.active
                
                # Leer filas (omitir la primera fila de encabezados)
                first_row = True
                for row in ws.rows:
                    if first_row:
                        first_row = False
                        continue
                    
                    if len(row) >= 2 and row[0].value and row[1].value:
                        rows.append([str(row[0].value), str(row[1].value)])
            except Exception as e:
                return templates.TemplateResponse("recarga_masiva.html", {
                    "request": request, 
                    "user": user,
                    "error": f"Error procesando archivo Excel: {str(e)}"
                })
        else:
            # Archivo CSV - intentar con diferentes codificaciones
            encodings = ['utf-8', 'latin-1', 'windows-1252', 'iso-8859-1']
            decoded = False
            
            for encoding in encodings:
                try:
                    # Intentar decodificar con esta codificación
                    text_content = content.decode(encoding).splitlines()
                    reader = csv.reader(text_content)
                    
                    # Omitir encabezados
                    next(reader, None)
                    
                    # Leer filas
                    for row in reader:
                        if len(row) >= 2:
                            rows.append(row)
                    
                    decoded = True
                    break  # Si llegamos aquí, la decodificación fue exitosa
                except UnicodeDecodeError:
                    continue  # Probar con la siguiente codificación
            
            if not decoded:
                return templates.TemplateResponse("recarga_masiva.html", {
                    "request": request, 
                    "user": user,
                    "error": "No se pudo decodificar el archivo. Asegúrese de que sea un CSV válido."
                })
        
        # Procesar las filas y actualizar la base de datos
        procesados = 0
        errores = []
        
        for row in rows:
            try:
                if len(row) >= 2:
                    # Extraer anexo y monto
                    calling_number = str(row[0]).strip()
                    
                    # Convertir monto a float, manejando posibles formatos
                    try:
                        amount = float(str(row[1]).strip().replace(',', '.'))
                    except ValueError:
                        errores.append(f"Error en anexo {calling_number}: '{row[1]}' no es un monto válido.")
                        continue
                    
                    # Buscar saldo actual
                    saldo_query = text("SELECT saldo FROM saldo_anexos WHERE calling_number = :calling_number")
                    saldo_actual = db.execute(saldo_query, {"calling_number": calling_number}).fetchone()

                    # Actualizar o insertar saldo
                    if saldo_actual:
                        update_query = text(
                            "UPDATE saldo_anexos SET saldo = saldo + :amount, fecha_ultima_recarga = CURRENT_TIMESTAMP WHERE calling_number = :calling_number"
                        )
                        db.execute(update_query, {"amount": amount, "calling_number": calling_number})
                    else:
                        insert_query = text(
                            "INSERT INTO saldo_anexos (calling_number, saldo, fecha_ultima_recarga) VALUES (:calling_number, :amount, CURRENT_TIMESTAMP)"
                        )
                        db.execute(insert_query, {"calling_number": calling_number, "amount": amount})

                    # Registrar en auditoría
                    audit_query = text(
                        "INSERT INTO saldo_auditoria (calling_number, saldo_anterior, saldo_nuevo, tipo_accion) "
                        "VALUES (:calling_number, :saldo_anterior, :saldo_nuevo, 'recarga_masiva')"
                    )
                    db.execute(audit_query, {
                        "calling_number": calling_number,
                        "saldo_anterior": saldo_actual[0] if saldo_actual else 0,
                        "saldo_nuevo": saldo_actual[0] + Decimal(str(amount)) if saldo_actual else Decimal(str(amount))
                    })
                    
                    procesados += 1
            except Exception as e:
                errores.append(f"Error procesando anexo {calling_number}: {str(e)}")
        
        # Confirmar transacción
        db.commit()
        
        # Generar mensaje de éxito
        success_message = f"Se procesaron {procesados} recargas exitosamente."
        
        # Devolver respuesta
        return templates.TemplateResponse("recarga_masiva.html", {
            "request": request, 
            "user": user,
            "success": success_message,
            "errores": errores if errores else None
        })
            
    except Exception as e:
        # Revertir cambios en caso de error
        db.rollback()
        
        return templates.TemplateResponse("recarga_masiva.html", {
            "request": request, 
            "user": user,
            "error": f"Error procesando el archivo: {str(e)}"
        })
    finally:
        db.close()

@app.get("/")
def root():
    return RedirectResponse(url="/dashboard/saldo")

@app.get("/logout")
def logout():
    response = RedirectResponse(url="/login")
    response.delete_cookie(manager.cookie_name)
    return response

# DASHBOARD: ANEXOS
@app.get("/dashboard/anexos")
async def dashboard_anexos(request: Request, 
                           user=Depends(admin_only),
                           page: int = Query(1, ge=1),
                           per_page: int = Query(20, ge=1, le=100),
                           buscar: str = Query(None)):
    if isinstance(user, RedirectResponse):
        return user
        
    db = SessionLocal()
    
    # Construir la consulta base para contar y para obtener registros
    base_query_str = """
        SELECT a.id, a.numero, a.usuario, a.area_nivel1, a.area_nivel2, a.area_nivel3, a.saldo_actual, a.pin, a.activo, s.saldo
        FROM anexos a
        LEFT JOIN saldo_anexos s ON a.numero = s.calling_number
        WHERE 1=1
    """
    count_query_str = "SELECT COUNT(*) FROM anexos WHERE 1=1"
    
    # Agregar condiciones de filtrado
    params = {}
    if buscar:
        base_query_str += """ 
            AND (a.numero LIKE :buscar OR a.usuario LIKE :buscar OR 
                 a.area_nivel1 LIKE :buscar OR a.area_nivel2 LIKE :buscar OR 
                 a.area_nivel3 LIKE :buscar)
        """
        count_query_str += """ 
            AND (numero LIKE :buscar OR usuario LIKE :buscar OR 
                 area_nivel1 LIKE :buscar OR area_nivel2 LIKE :buscar OR 
                 area_nivel3 LIKE :buscar)
        """
        params["buscar"] = f"%{buscar}%"

    # Calcular el total de registros para la paginación
    count_query = text(count_query_str)
    total_records = db.execute(count_query, params).scalar()
    total_pages = (total_records + per_page - 1) // per_page  # Ceiling division
    
    # Agregar límite y offset para la paginación
    offset = (page - 1) * per_page
    base_query_str += f" ORDER BY a.numero ASC LIMIT {per_page} OFFSET {offset}"
    
    # Ejecutar la consulta principal
    query = text(base_query_str)
    rows = db.execute(query, params).fetchall()
    
    # Consultar el valor de longitud de PIN actual
    pin_length_query = text("SELECT valor FROM configuracion WHERE clave = 'pin_length'")
    pin_length_row = db.execute(pin_length_query).fetchone()
    pin_length = int(pin_length_row[0]) if pin_length_row else 6  # Valor por defecto: 6
    
    # Consultar áreas para los dropdowns
    areas_nivel1_query = text("SELECT DISTINCT area_nivel1 FROM anexos WHERE area_nivel1 IS NOT NULL ORDER BY area_nivel1")
    areas_nivel1 = [row[0] for row in db.execute(areas_nivel1_query).fetchall()]
    
    areas_nivel2_query = text("SELECT DISTINCT area_nivel2 FROM anexos WHERE area_nivel2 IS NOT NULL ORDER BY area_nivel2")
    areas_nivel2 = [row[0] for row in db.execute(areas_nivel2_query).fetchall()]
    
    areas_nivel3_query = text("SELECT DISTINCT area_nivel3 FROM anexos WHERE area_nivel3 IS NOT NULL ORDER BY area_nivel3")
    areas_nivel3 = [row[0] for row in db.execute(areas_nivel3_query).fetchall()]
    
    db.close()

    return templates.TemplateResponse("dashboard_anexos.html", {
        "request": request,
        "rows": rows,
        "page": page,
        "per_page": per_page,
        "total_pages": total_pages,
        "total_records": total_records,
        "buscar": buscar or "",
        "user": user,
        "pin_length": pin_length,
        "areas_nivel1": areas_nivel1,
        "areas_nivel2": areas_nivel2,
        "areas_nivel3": areas_nivel3
    })

@app.get("/anexo/{anexo_id}")
async def get_anexo(anexo_id: int, user=Depends(admin_only)):
    if isinstance(user, RedirectResponse):
        return user
        
    db = SessionLocal()
    query = text("SELECT id, numero, usuario, area_nivel1, area_nivel2, area_nivel3, saldo_actual, activo FROM anexos WHERE id = :anexo_id")
    anexo = db.execute(query, {"anexo_id": anexo_id}).fetchone()
    db.close()
    
    if not anexo:
        raise HTTPException(status_code=404, detail="Anexo no encontrado")
        
    return {
        "id": anexo[0],
        "numero": anexo[1],
        "usuario": anexo[2],
        "area_nivel1": anexo[3],
        "area_nivel2": anexo[4],
        "area_nivel3": anexo[5],
        "saldo_actual": float(anexo[6]),
        "activo": anexo[7]
    }

# Modelo para crear/actualizar anexo
class AnexoCreate(BaseModel):
    numero: str
    usuario: str
    area_nivel1: str
    area_nivel2: Optional[str] = None
    area_nivel3: Optional[str] = None
    pin: Optional[str] = None
    saldo_actual: Optional[float] = 0
    activo: Optional[bool] = True
    
    # Validadores (opcional)
    @validator('numero')
    def numero_valid(cls, v):
        if not v or not v.strip():
            raise ValueError('El número no puede estar vacío')
        return v
    
    @validator('usuario')
    def usuario_valid(cls, v):
        if not v or not v.strip():
            raise ValueError('El usuario no puede estar vacío')
        return v
    
    #@validator('saldo_actual')
    #def saldo_valid(cls, v):
    #    if v < 0:
    #        raise ValueError('El saldo inicial no puede ser negativo')
    #    return v


@app.post("/anexo")
async def crear_anexo(anexo: AnexoCreate, user=Depends(admin_only)):
    if isinstance(user, RedirectResponse):
        return user
    
    # Imprimir toda la información para depuración
    print(f"Datos recibidos del formulario:")
    print(f"  Número: '{anexo.numero}'")
    print(f"  Usuario: '{anexo.usuario}'")
    print(f"  Área Nivel 1: '{anexo.area_nivel1}'")
    print(f"  Área Nivel 2: '{getattr(anexo, 'area_nivel2', '')}'")
    print(f"  Área Nivel 3: '{getattr(anexo, 'area_nivel3', '')}'")
    print(f"  PIN: '{'*****' if anexo.pin else 'No proporcionado'}'")
    #print(f"  Saldo Actual: {anexo.saldo_actual}")
    print(f"  Activo: {getattr(anexo, 'activo', True)}")
            
    try:    
        db = SessionLocal()
        
        # Verificar si el número de anexo ya existe
        check_query = text("SELECT id FROM anexos WHERE numero = :numero")
        existing = db.execute(check_query, {"numero": anexo.numero}).fetchone()
        
        if existing:
            print(f"Error: El anexo {anexo.numero} ya existe")
            db.close()
            raise HTTPException(status_code=422, detail=f"El número de anexo {anexo.numero} ya existe")
        
        # Generar PIN automático si no se proporciona
        pin = anexo.pin
        if not pin:
            # Obtener longitud configurada del PIN
            try:
                pin_length_query = text("SELECT valor FROM configuracion WHERE clave = 'pin_length'")
                pin_length_row = db.execute(pin_length_query).fetchone()
                pin_length = int(pin_length_row[0]) if pin_length_row else 6
            except Exception as e:
                print(f"Error al obtener longitud de PIN: {e}")
                pin_length = 6  # Valor por defecto si hay error
            
            # Generar PIN aleatorio
            import random
            pin = ''.join(random.choices('0123456789', k=pin_length))
            print(f"PIN generado automáticamente: {pin}")
        
        # Hashear el PIN para almacenarlo de forma segura
        #try:
        #    hashed_pin = pwd_context.hash(pin)
        #except Exception as e:
        #    print(f"Error al hashear PIN: {e}")
        #    raise HTTPException(status_code=500, detail="Error al procesar el PIN")
        
        # Preparar los parámetros para la inserción, garantizando valores por defecto para campos opcionales
        params = {
            "numero": anexo.numero,
            "usuario": anexo.usuario,
            "area_nivel1": anexo.area_nivel1,
            "area_nivel2": anexo.area_nivel2 if hasattr(anexo, 'area_nivel2') and anexo.area_nivel2 else "",
            "area_nivel3": anexo.area_nivel3 if hasattr(anexo, 'area_nivel3') and anexo.area_nivel3 else "",
            "pin": pin,
            "activo": anexo.activo if hasattr(anexo, 'activo') else True
        }
        
        # Insertar el nuevo anexo
        try:
            insert_query = text("""
                INSERT INTO anexos (numero, usuario, area_nivel1, area_nivel2, area_nivel3, pin, activo)
                VALUES (:numero, :usuario, :area_nivel1, :area_nivel2, :area_nivel3, :pin, :activo)
                RETURNING id
            """)
            
            result = db.execute(insert_query, params)
            anexo_id = result.fetchone()[0]
            print(f"Anexo creado con ID: {anexo_id}")
            
        except Exception as e:
            print(f"Error al insertar anexo: {e}")
            db.rollback()
            db.close()
            raise HTTPException(status_code=500, detail=f"Error al crear anexo: {str(e)}")
        
        # Inicializar saldo en la tabla saldo_anexos si tiene saldo inicial
        try:
            if anexo.saldo_actual > 0:
                saldo_query = text("""
                    INSERT INTO saldo_anexos (calling_number, saldo, fecha_ultima_recarga)
                    VALUES (:numero, :saldo, CURRENT_TIMESTAMP)
                """)
                db.execute(saldo_query, {"numero": anexo.numero, "saldo": anexo.saldo_actual})
                
                # Registrar en auditoría
                audit_query = text("""
                    INSERT INTO saldo_auditoria (calling_number, saldo_anterior, saldo_nuevo, tipo_accion)
                    VALUES (:numero, 0, :saldo, 'creacion_anexo')
                """)
                db.execute(audit_query, {"numero": anexo.numero, "saldo": anexo.saldo_actual})
                
                print(f"Saldo inicial registrado: {anexo.saldo_actual}")
        except Exception as e:
            print(f"Error al registrar saldo: {e}")
            # No fallamos la operación completa si solo falla la parte del saldo
            # Pero registramos el error para investigación
        
        db.commit()
        db.close()
        
        return {"id": anexo_id, "mensaje": "Anexo creado exitosamente", "pin": pin}
    
    except HTTPException:
        # Re-lanzar excepciones HTTP tal cual
        raise
    
    except Exception as e:
        print(f"Error inesperado creando anexo: {e}")
        # Si llegamos aquí es un error inesperado
        raise HTTPException(status_code=500, detail=f"Error interno del servidor: {str(e)}")

    
@app.put("/anexo/{anexo_id}")
async def actualizar_anexo(anexo_id: int, anexo: AnexoCreate, user=Depends(admin_only)):
    if isinstance(user, RedirectResponse):
        return user
        
    db = SessionLocal()
    
    # Verificar si el anexo existe
    check_query = text("SELECT numero, saldo_actual FROM anexos WHERE id = :anexo_id")
    existing = db.execute(check_query, {"anexo_id": anexo_id}).fetchone()
    
    if not existing:
        db.close()
        raise HTTPException(status_code=404, detail="Anexo no encontrado")
    
    old_numero = existing[0]
    old_saldo = float(existing[1])
    
    # Verificar si el nuevo número ya existe (y no es el mismo anexo)
    if anexo.numero != old_numero:
        check_duplicate_query = text("SELECT id FROM anexos WHERE numero = :numero")
        duplicate = db.execute(check_duplicate_query, {"numero": anexo.numero}).fetchone()
        
        if duplicate:
            db.close()
            raise HTTPException(status_code=400, detail="El número de anexo ya existe")
    
    # Preparar actualización de PIN si se proporciona
    params = {
        "anexo_id": anexo_id,
        "numero": anexo.numero,
        "usuario": anexo.usuario,
        "area_nivel1": anexo.area_nivel1,
        "area_nivel2": anexo.area_nivel2,
        "area_nivel3": anexo.area_nivel3,
        "activo": anexo.activo
    }
    
    update_query_str = """
        UPDATE anexos SET 
        numero = :numero,
        usuario = :usuario,
        area_nivel1 = :area_nivel1,
        area_nivel2 = :area_nivel2,
        area_nivel3 = :area_nivel3,
        activo = :activo
    """
    
    # Si se proporciona PIN, actualizarlo
    if anexo.pin:
        #hashed_pin = pwd_context.hash(anexo.pin)
        update_query_str += ", pin = :pin"
        params["pin"] = anexo.pin
    
    update_query_str += " WHERE id = :anexo_id"
    update_query = text(update_query_str)
    
    db.execute(update_query, params)
    
    db.commit()
    db.close()
    
    return {"mensaje": "Anexo actualizado exitosamente"}

@app.delete("/anexo/{anexo_id}")
async def eliminar_anexo(anexo_id: int, user=Depends(admin_only)):
    if isinstance(user, RedirectResponse):
        return user
        
    db = SessionLocal()
    
    try:
        # Verificar si el anexo existe
        check_query = text("SELECT numero FROM anexos WHERE id = :anexo_id")
        existing = db.execute(check_query, {"anexo_id": anexo_id}).fetchone()
        
        if not existing:
            raise HTTPException(status_code=404, detail="Anexo no encontrado")
        
        numero_anexo = existing[0]
        
        # Eliminación física (delete del registro)
        delete_query = text("DELETE FROM anexos WHERE id = :anexo_id")
        result = db.execute(delete_query, {"anexo_id": anexo_id})
        
        # Verificar que se eliminó efectivamente
        if result.rowcount == 0:
            raise HTTPException(status_code=500, detail="Error al eliminar el anexo")
        
        db.commit()
        
        return {"mensaje": f"Anexo {numero_anexo} eliminado exitosamente"}
        
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error interno del servidor: {str(e)}")
    finally:
        db.close()

# Configuración de longitud de PIN
@app.put("/configuracion/pin_length")
async def actualizar_longitud_pin(pin_length: int = Form(...), user=Depends(admin_only)):
    if isinstance(user, RedirectResponse):
        return user
        
    if pin_length < 4 or pin_length > 10:
        raise HTTPException(status_code=400, detail="La longitud del PIN debe estar entre 4 y 10 dígitos")
        
    db = SessionLocal()
    
    # Verificar si ya existe la configuración
    check_query = text("SELECT id FROM configuracion WHERE clave = 'pin_length'")
    existing = db.execute(check_query).fetchone()
    
    if existing:
        # Actualizar configuración existente
        update_query = text("UPDATE configuracion SET valor = :valor WHERE clave = 'pin_length'")
        db.execute(update_query, {"valor": str(pin_length)})
    else:
        # Crear nueva configuración
        insert_query = text("""
            INSERT INTO configuracion (clave, valor, descripcion)
            VALUES ('pin_length', :valor, 'Longitud de PIN para anexos')
        """)
        db.execute(insert_query, {"valor": str(pin_length)})
    
    db.commit()
    db.close()
    
    return {"mensaje": f"Longitud de PIN actualizada a {pin_length} dígitos"}

# Generación masiva de PINs
@app.post("/anexos/generar_pines")
async def generar_pines_masivos(request: Request, user=Depends(admin_only)):
    if isinstance(user, RedirectResponse):
        return user
        
    db = SessionLocal()
    
    # Obtener longitud configurada del PIN
    pin_length_query = text("SELECT valor FROM configuracion WHERE clave = 'pin_length'")
    pin_length_row = db.execute(pin_length_query).fetchone()
    pin_length = int(pin_length_row[0]) if pin_length_row else 6
    
    # Obtener todos los anexos activos
    anexos_query = text("SELECT id, numero FROM anexos WHERE activo = TRUE")
    anexos = db.execute(anexos_query).fetchall()
    
    import random
    resultados = []
    
    for anexo in anexos:
        anexo_id, numero = anexo
        
        # Generar nuevo PIN
        pin = ''.join(random.choices('0123456789', k=pin_length))
        hashed_pin = pwd_context.hash(pin)
        
        # Actualizar PIN
        update_query = text("UPDATE anexos SET pin = :pin WHERE id = :anexo_id")
        db.execute(update_query, {"pin": hashed_pin, "anexo_id": anexo_id})
        
        resultados.append({"numero": numero, "pin": pin})
    
    db.commit()
    db.close()
    
    # Generar CSV para descarga
    csv_content = io.StringIO()
    writer = csv.writer(csv_content)
    writer.writerow(["Número de Anexo", "PIN"])
    
    for resultado in resultados:
        writer.writerow([resultado["numero"], resultado["pin"]])
    
    csv_content.seek(0)
    
    return StreamingResponse(
        io.BytesIO(csv_content.getvalue().encode()),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=pines_anexos_{datetime.now().strftime('%Y%m%d%H%M%S')}.csv"}
    )

# Conexion CUCM
@app.get("/dashboard/cucm")
async def dashboard_cucm(request: Request, user=Depends(admin_only)):
    if isinstance(user, RedirectResponse):
        return user
        
    return templates.TemplateResponse("dashboard_cucm.html", {
        "request": request,
        "user": user
    })

# Carga masiva de anexos
@app.get("/dashboard/anexos/carga_masiva")
async def form_carga_masiva_anexos(request: Request, user=Depends(admin_only)):
    if isinstance(user, RedirectResponse):
        return user
        
    return templates.TemplateResponse("carga_masiva_anexos.html", {"request": request, "user": user})

@app.post("/dashboard/anexos/carga_masiva")
async def carga_masiva_anexos(
    request: Request, 
    file: UploadFile = File(...), 
    generar_pin: bool = Form(True),
    continuar_errores: bool = Form(False),
    user=Depends(admin_only)
):
    """
    Procesa un archivo CSV o Excel para cargar anexos masivamente.
    
    Params:
    - file: Archivo CSV o Excel con los datos de anexos
    - generar_pin: Si se debe generar PIN automáticamente cuando no se proporciona
    - continuar_errores: Si se debe continuar procesando a pesar de errores
    """
    if isinstance(user, RedirectResponse):
        return user
    
    db = SessionLocal()
    
    # Variables para el seguimiento
    procesados = 0
    exitosos = 0
    errores = []
    
    try:
        # Leer el contenido del archivo
        content = await file.read()
        
        # Determinar tipo de archivo por extensión
        filename = file.filename.lower()
        
        # Lista para almacenar las filas a procesar
        rows = []
        headers = []
        
        # Procesar según el tipo de archivo
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            # Procesar archivo Excel
            try:
                # Intentar usar la biblioteca openpyxl si está disponible
                try:
                    import io
                    import openpyxl
                    
                    # Cargar el archivo Excel
                    wb = openpyxl.load_workbook(io.BytesIO(content))
                    ws = wb.active
                    
                    # Leer encabezados (primera fila)
                    headers = [str(cell.value).strip() if cell.value else "" for cell in next(ws.rows)]
                    
                    # Verificar encabezados mínimos requeridos
                    if not headers or 'numero' not in headers or 'usuario' not in headers or 'area_nivel1' not in headers:
                        return templates.TemplateResponse("carga_masiva_anexos.html", {
                            "request": request, 
                            "user": user,
                            "error": "El archivo no tiene los encabezados requeridos: numero, usuario, area_nivel1"
                        })
                    
                    # Leer filas (omitir la primera fila de encabezados)
                    for row in list(ws.rows)[1:]:
                        # Convertir celdas a valores de texto
                        row_values = [str(cell.value).strip() if cell.value is not None else "" for cell in row]
                        if any(row_values):  # Omitir filas vacías
                            rows.append(row_values)
                    
                except ImportError:
                    # Si openpyxl no está disponible, intentar usar xlrd (para archivos .xls)
                    try:
                        import io
                        import xlrd
                        
                        # Abrir el libro Excel
                        workbook = xlrd.open_workbook(file_contents=content)
                        sheet = workbook.sheet_by_index(0)
                        
                        # Leer encabezados (primera fila)
                        headers = [str(sheet.cell_value(0, col)).strip() for col in range(sheet.ncols)]
                        
                        # Verificar encabezados mínimos requeridos
                        if not headers or 'numero' not in headers or 'usuario' not in headers or 'area_nivel1' not in headers:
                            return templates.TemplateResponse("carga_masiva_anexos.html", {
                                "request": request, 
                                "user": user,
                                "error": "El archivo no tiene los encabezados requeridos: numero, usuario, area_nivel1"
                            })
                        
                        # Leer filas (omitir la primera fila de encabezados)
                        for row_idx in range(1, sheet.nrows):
                            row_values = [str(sheet.cell_value(row_idx, col)).strip() for col in range(sheet.ncols)]
                            if any(row_values):  # Omitir filas vacías
                                rows.append(row_values)
                    
                    except ImportError:
                        # Si ninguna biblioteca Excel está disponible, intentar con pandas
                        try:
                            import io
                            import pandas as pd
                            
                            # Leer el archivo Excel con pandas
                            df = pd.read_excel(io.BytesIO(content))
                            
                            # Verificar encabezados mínimos requeridos
                            df_columns = df.columns.tolist()
                            headers = [str(col).strip() for col in df_columns]
                            
                            if not headers or 'numero' not in headers or 'usuario' not in headers or 'area_nivel1' not in headers:
                                return templates.TemplateResponse("carga_masiva_anexos.html", {
                                    "request": request, 
                                    "user": user,
                                    "error": "El archivo no tiene los encabezados requeridos: numero, usuario, area_nivel1"
                                })
                            
                            # Convertir DataFrame a lista de filas
                            rows = df.values.tolist()
                            
                        except ImportError:
                            # Si ninguna de las bibliotecas está disponible
                            return templates.TemplateResponse("carga_masiva_anexos.html", {
                                "request": request, 
                                "user": user,
                                "error": "No se pueden procesar archivos Excel en este servidor. Por favor, exporte a CSV e intente de nuevo."
                            })
            
            except Exception as e:
                print(f"Error procesando archivo Excel: {str(e)}")
                return templates.TemplateResponse("carga_masiva_anexos.html", {
                    "request": request, 
                    "user": user,
                    "error": f"Error procesando archivo Excel: {str(e)}"
                })
        
        else:
            # Archivo CSV - intentar con diferentes codificaciones
            encodings = ['utf-8', 'latin-1', 'windows-1252', 'iso-8859-1']
            decoded = False
            
            for encoding in encodings:
                try:
                    # Intentar decodificar con esta codificación
                    text_content = content.decode(encoding).splitlines()
                    reader = csv.reader(text_content)
                    
                    # Leer encabezados
                    headers = next(reader, None)
                    
                    # Verificar encabezados mínimos requeridos
                    if not headers or 'numero' not in headers or 'usuario' not in headers or 'area_nivel1' not in headers:
                        return templates.TemplateResponse("carga_masiva_anexos.html", {
                            "request": request, 
                            "user": user,
                            "error": "El archivo no tiene los encabezados requeridos: numero, usuario, area_nivel1"
                        })
                    
                    # Leer filas
                    rows = list(reader)
                    decoded = True
                    print(f"✅ Archivo decodificado correctamente con codificación: {encoding}")
                    break  # Si llegamos aquí, la decodificación fue exitosa
                except UnicodeDecodeError:
                    print(f"❌ Fallo al decodificar con {encoding}")
                    continue  # Probar con la siguiente codificación
            
            if not decoded:
                return templates.TemplateResponse("carga_masiva_anexos.html", {
                    "request": request, 
                    "user": user,
                    "error": "No se pudo decodificar el archivo. Asegúrese de que sea un CSV válido."
                })
        
        # Si no hay filas para procesar
        if not rows:
            return templates.TemplateResponse("carga_masiva_anexos.html", {
                "request": request, 
                "user": user,
                "error": "El archivo no contiene datos para procesar."
            })
        
        # Mapear índices de columnas
        column_indices = {}
        for col in ['numero', 'usuario', 'area_nivel1', 'area_nivel2', 'area_nivel3', 'pin', 'saldo_actual']:
            column_indices[col] = headers.index(col) if col in headers else -1
        
        # Obtener longitud configurada del PIN
        pin_length_query = text("SELECT valor FROM configuracion WHERE clave = 'pin_length'")
        pin_length_row = db.execute(pin_length_query).fetchone()
        pin_length = int(pin_length_row[0]) if pin_length_row else 6  # Valor por defecto: 6
        
        # Procesar filas
        for i, row in enumerate(rows, start=2):  # start=2 para considerar la fila 1 como encabezados
            try:
                # Asegurarse de que la fila tiene suficientes columnas
                if len(row) <= max(idx for idx in column_indices.values() if idx >= 0):
                    # Extender la fila con valores vacíos si es necesario
                    row.extend([''] * (max(idx for idx in column_indices.values() if idx >= 0) - len(row) + 1))
                
                # Validar que tenga los campos requeridos
                if column_indices['numero'] < 0 or column_indices['usuario'] < 0 or column_indices['area_nivel1'] < 0:
                    raise ValueError("No se encontraron las columnas requeridas: numero, usuario, area_nivel1")
                
                # Verificar que los campos requeridos no estén vacíos
                if not row[column_indices['numero']] or not row[column_indices['usuario']] or not row[column_indices['area_nivel1']]:
                    raise ValueError("Faltan valores en campos requeridos: numero, usuario, area_nivel1")
                
                # Extraer datos básicos
                numero = str(row[column_indices['numero']]).strip()
                usuario = str(row[column_indices['usuario']]).strip()
                area_nivel1 = str(row[column_indices['area_nivel1']]).strip()
                
                # Extraer datos opcionales
                area_nivel2 = str(row[column_indices['area_nivel2']]).strip() if column_indices['area_nivel2'] >= 0 and len(row) > column_indices['area_nivel2'] else ''
                area_nivel3 = str(row[column_indices['area_nivel3']]).strip() if column_indices['area_nivel3'] >= 0 and len(row) > column_indices['area_nivel3'] else ''
                
                # Extraer PIN si existe
                pin = None
                if column_indices['pin'] >= 0 and len(row) > column_indices['pin'] and row[column_indices['pin']]:
                    pin = str(row[column_indices['pin']]).strip()
                
                # Extraer saldo inicial si existe
                saldo_actual = 0.0
                if column_indices['saldo_actual'] >= 0 and len(row) > column_indices['saldo_actual'] and row[column_indices['saldo_actual']]:
                    try:
                        saldo_value = str(row[column_indices['saldo_actual']]).strip().replace(',', '.')
                        # Manejar casos donde el valor puede ser un float en formato de texto o tener caracteres no numéricos
                        saldo_actual = float(''.join(c for c in saldo_value if c.isdigit() or c == '.'))
                    except ValueError:
                        print(f"⚠️ Error convirtiendo saldo: {row[column_indices['saldo_actual']]}")
                        saldo_actual = 0.0
                
                # Verificar si el número ya existe
                check_query = text("SELECT id FROM anexos WHERE numero = :numero")
                existing = db.execute(check_query, {"numero": numero}).fetchone()
                
                if existing:
                    raise ValueError(f"El anexo {numero} ya existe")
                
                # Generar PIN si no se proporcionó y está activada la opción
                if not pin and generar_pin:
                    import random
                    pin = ''.join(random.choices('0123456789', k=pin_length))
                    print(f"PIN generado para {numero}: {pin}")
                
                # Hashear el PIN
                from passlib.context import CryptContext
                pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
                hashed_pin = pwd_context.hash(pin) if pin else None
                
                # Insertar el nuevo anexo
                insert_query = text("""
                    INSERT INTO anexos (numero, usuario, area_nivel1, area_nivel2, area_nivel3, pin, saldo_actual, activo)
                    VALUES (:numero, :usuario, :area_nivel1, :area_nivel2, :area_nivel3, :pin, :saldo_actual, TRUE)
                    RETURNING id
                """)
                
                result = db.execute(insert_query, {
                    "numero": numero,
                    "usuario": usuario,
                    "area_nivel1": area_nivel1,
                    "area_nivel2": area_nivel2,
                    "area_nivel3": area_nivel3,
                    "pin": hashed_pin,
                    "saldo_actual": saldo_actual,
                    "activo": True
                })
                
                anexo_id = result.fetchone()[0]
                
                # Inicializar saldo en la tabla saldo_anexos si tiene saldo inicial
                if saldo_actual > 0:
                    saldo_query = text("""
                        INSERT INTO saldo_anexos (calling_number, saldo, fecha_ultima_recarga)
                        VALUES (:numero, :saldo, CURRENT_TIMESTAMP)
                    """)
                    db.execute(saldo_query, {"numero": numero, "saldo": saldo_actual})
                    
                    # Registrar en auditoría
                    audit_query = text("""
                        INSERT INTO saldo_auditoria (calling_number, saldo_anterior, saldo_nuevo, tipo_accion)
                        VALUES (:numero, 0, :saldo, 'creacion_anexo')
                    """)
                    db.execute(audit_query, {"numero": numero, "saldo": saldo_actual})
                
                procesados += 1
                exitosos += 1
                
            except Exception as e:
                # Registrar error
                error_msg = f"Fila {i}: {str(e)}"
                errores.append(error_msg)
                
                # Si no debemos continuar ante errores, terminar
                if not continuar_errores:
                    db.rollback()
                    return templates.TemplateResponse("carga_masiva_anexos.html", {
                        "request": request,
                        "user": user,
                        "error": f"Error en la fila {i}: {str(e)}. Proceso abortado.",
                        "errores": errores
                    })
        
        # Confirmar cambios en la base de datos
        db.commit()
        
        # Mensaje de éxito
        success_message = f"Se procesaron {procesados} registros. {exitosos} anexos creados con éxito."
        if errores:
            success_message += f" Se encontraron {len(errores)} errores."
        
        return templates.TemplateResponse("carga_masiva_anexos.html", {
            "request": request,
            "user": user,
            "success": success_message,
            "errores": errores if errores else None
        })
        
    except Exception as e:
        # En caso de error general, hacer rollback
        db.rollback()
        return templates.TemplateResponse("carga_masiva_anexos.html", {
            "request": request,
            "user": user,
            "error": f"Error general: {str(e)}"
        })
    finally:
        db.close()
        
# Modelo Pydantic para la configuración de CUCM
class CucmConfigModel(BaseModel):
    server_ip: str
    server_port: int = 2748
    username: str
    password: str
    app_info: str = "TarificadorApp"
    reconnect_delay: int = 30
    check_interval: int = 60
    enabled: bool = True

# Ruta para obtener la configuración actual
@app.get("/api/cucm/config")
async def get_cucm_config(user=Depends(admin_only)):
    if isinstance(user, RedirectResponse):
        return user
        
    db = SessionLocal()
    query = text("SELECT server_ip, server_port, username, password, app_info, reconnect_delay, check_interval, enabled, last_status, last_status_update FROM cucm_config ORDER BY id DESC LIMIT 1")
    result = db.execute(query).fetchone()
    
    if not result:
        # Si no hay configuración, devolver valores por defecto
        db.close()
        return {
            "server_ip": "10.224.0.10",
            "server_port": 2748,
            "username": "jtapiuser",
            "password": "********",  # Ocultamos la contraseña real
            "app_info": "TarificadorApp",
            "reconnect_delay": 30,
            "check_interval": 60,
            "enabled": True,
            "last_status": "unknown",
            "last_status_update": None
        }
    
    config = {
        "server_ip": result[0],
        "server_port": result[1],
        "username": result[2],
        "password": "********",  # Ocultamos la contraseña real
        "app_info": result[4],
        "reconnect_delay": result[5],
        "check_interval": result[6],
        "enabled": result[7],
        "last_status": result[8],
        "last_status_update": result[9]
    }
    
    db.close()
    return config

# Ruta para actualizar la configuración de CUCM
@app.post("/api/cucm/config")
async def update_cucm_config(config: CucmConfigModel, user=Depends(admin_only)):
    if isinstance(user, RedirectResponse):
        return user
        
    db = SessionLocal()
    
    # Verificar si ya existe una configuración
    check_query = text("SELECT id FROM cucm_config LIMIT 1")
    existing = db.execute(check_query).fetchone()
    
    if existing:
        # Actualizar configuración existente
        update_query = text("""
            UPDATE cucm_config SET 
            server_ip = :server_ip,
            server_port = :server_port,
            username = :username,
            password = :password,
            app_info = :app_info,
            reconnect_delay = :reconnect_delay,
            check_interval = :check_interval,
            enabled = :enabled,
            last_updated = CURRENT_TIMESTAMP
            WHERE id = :id
        """)
        
        db.execute(update_query, {
            "id": existing[0],
            "server_ip": config.server_ip,
            "server_port": config.server_port,
            "username": config.username,
            "password": config.password,
            "app_info": config.app_info,
            "reconnect_delay": config.reconnect_delay,
            "check_interval": config.check_interval,
            "enabled": config.enabled
        })
    else:
        # Crear nueva configuración
        insert_query = text("""
            INSERT INTO cucm_config (server_ip, server_port, username, password, app_info, reconnect_delay, check_interval, enabled)
            VALUES (:server_ip, :server_port, :username, :password, :app_info, :reconnect_delay, :check_interval, :enabled)
        """)
        
        db.execute(insert_query, {
            "server_ip": config.server_ip,
            "server_port": config.server_port,
            "username": config.username,
            "password": config.password,
            "app_info": config.app_info,
            "reconnect_delay": config.reconnect_delay,
            "check_interval": config.check_interval,
            "enabled": config.enabled
        })
    
    # Generar nuevo archivo de configuración para el servicio Java
    generate_java_config(config, db)
    
    # Señalizar al servicio que debe recargar la configuración
    notify_java_service(config)
    
    db.commit()
    db.close()
    
    return {"message": "Configuración actualizada correctamente"}

# Ruta para probar la conexión
@app.post("/api/cucm/test_connection")
async def test_cucm_connection(config: CucmConfigModel, user=Depends(admin_only)):
    if isinstance(user, RedirectResponse):
        return user
    
    # Aquí implementamos la lógica para probar la conexión
    # Esta función debería llamar al servicio Java para probar la conexión
    # y devolver el resultado

    try:
        # Simulamos un retraso de respuesta del servicio
        await asyncio.sleep(2)
        
        # En un entorno real, esto llamaría a un endpoint en el servicio Java
        # Por ahora, simplemente simularemos una respuesta exitosa
        success = True
        message = "Conexión exitosa con el servidor CUCM"
        
        # Si la conexión es exitosa, actualizamos el estado en la base de datos
        if success:
            db = SessionLocal()
            update_query = text("""
                UPDATE cucm_config SET 
                last_status = :status,
                last_status_update = CURRENT_TIMESTAMP
                WHERE id = (SELECT id FROM cucm_config ORDER BY id DESC LIMIT 1)
            """)
            
            db.execute(update_query, {"status": "connected"})
            db.commit()
            db.close()
        
        return {"success": success, "message": message}
    except Exception as e:
        return {"success": False, "message": f"Error al probar la conexión: {str(e)}"}

# Función para generar archivo de configuración Java
def generate_java_config(config, db):
    try:
        # Ruta del archivo de configuración
        config_path = "/opt/tarificador/java-service/config/application.properties"
        
        # Crear contenido del archivo
        content = f"""# Configuración generada automáticamente por el dashboard
# Última actualización: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

# Configuración de conexión CUCM
cucm.server.ip={config.server_ip}
cucm.server.port={config.server_port}
cucm.server.username={config.username}
cucm.server.password={config.password}
cucm.server.appinfo={config.app_info}
cucm.enabled={str(config.enabled).lower()}

# Configuración del servicio web
webservice.base.url=http://localhost:8000
webservice.endpoint.cdr=/cdr
webservice.endpoint.check_balance=/check_balance/
webservice.endpoint.call_start=/call_start

# Configuración de reconexión
reconnect.initial.delay={config.reconnect_delay}
reconnect.check.interval={config.check_interval}

# Configuración de logging
logging.level=INFO
logging.file.path=/var/log/tarificador
logging.file.name=cucm-service.log
"""
        
        # Escribir archivo
        with open(config_path, "w") as f:
            f.write(content)
            
        # Actualizar estado en la base de datos
        update_query = text("""
            UPDATE cucm_config SET 
            last_status = :status,
            last_status_update = CURRENT_TIMESTAMP
            WHERE id = (SELECT id FROM cucm_config ORDER BY id DESC LIMIT 1)
        """)
        
        db.execute(update_query, {"status": "config_updated"})
        
        return True
    except Exception as e:
        print(f"Error al generar archivo de configuración: {str(e)}")
        return False

# Función para notificar al servicio Java
def notify_java_service(config):
    try:
        # En un entorno real, esto podría:
        # 1. Enviar una señal al proceso Java (por ejemplo, usando SIGHUP)
        # 2. Usar un endpoint REST en el servicio Java para indicarle que recargue la configuración
        # 3. Usar un archivo temporal o un socket para comunicarse con el servicio
        
        # Por simplicidad, aquí usaremos systemctl para reiniciar el servicio
        if os.path.exists("/bin/systemctl"):
            if config.enabled:
                os.system("systemctl restart tarificador-cucm.service")
            else:
                os.system("systemctl stop tarificador-cucm.service")
        
        return True
    except Exception as e:
        print(f"Error al notificar al servicio Java: {str(e)}")
        return False


# Ruta para obtener el estado del servicio CUCM
@app.get("/api/cucm/status")
async def get_cucm_status(user=Depends(authenticated_user)):
    if isinstance(user, RedirectResponse):
        return user
        
    try:
        # Verificar el estado del servicio systemd
        is_active = False
        status_message = "Desconocido"
        
        if os.path.exists("/bin/systemctl"):
            # Verificar si el servicio está activo
            result = os.popen("systemctl is-active tarificador-cucm.service").read().strip()
            is_active = (result == "active")
            
            # Obtener estado más detallado
            status_output = os.popen("systemctl status tarificador-cucm.service --no-pager").read()
            
            # Extraer línea de estado
            if "Active:" in status_output:
                status_line = [line for line in status_output.split('\n') if "Active:" in line][0]
                status_message = status_line.strip()
        
        # Obtener el último estado registrado en la base de datos
        db = SessionLocal()
        query = text("SELECT last_status, last_status_update FROM cucm_config ORDER BY id DESC LIMIT 1")
        result = db.execute(query).fetchone()
        db.close()
        
        db_status = "unknown"
        db_status_time = None
        
        if result:
            db_status = result[0]
            db_status_time = result[1]
        
        return {
            "service_active": is_active,
            "service_status": status_message,
            "last_known_status": db_status,
            "last_status_update": db_status_time
        }
    except Exception as e:
        return {"error": f"Error al obtener estado: {str(e)}"}
    

# Endpoint para obtener logs del servicio
@app.get("/api/cucm/logs")
async def get_cucm_logs(lines: int = 100, user=Depends(authenticated_user)):
    if isinstance(user, RedirectResponse):
        return user
        
    try:
        log_path = "/var/log/tarificador/cucm-service.log"
        
        if not os.path.exists(log_path):
            return {"logs": "Archivo de log no encontrado"}
        
        # Leer las últimas líneas del archivo de log
        with open(log_path, 'r') as f:
            # Leer todo el archivo si es pequeño, o usar readlines con un buffer para archivos grandes
            if os.path.getsize(log_path) < 1024 * 1024:  # < 1MB
                all_lines = f.readlines()
            else:
                # Para archivos grandes, leer solo las últimas N líneas
                f.seek(0, os.SEEK_END)
                buffer_size = 8192
                file_size = f.tell()
                block_end = file_size
                
                # Colectar líneas hasta que tengamos suficientes o lleguemos al inicio del archivo
                all_lines = []
                while len(all_lines) < lines and block_end > 0:
                    block_start = max(0, block_end - buffer_size)
                    f.seek(block_start)
                    
                    # Si no estamos al inicio del archivo, descartar la primera línea parcial
                    if block_start > 0:
                        f.readline()
                    
                    # Leer las líneas del bloque
                    block_lines = f.readlines()
                    
                    # Actualizar para el siguiente bloque
                    block_end = block_start
                    
                    # Agregar líneas al inicio (para mantener el orden)
                    all_lines = block_lines + all_lines
                
                # Limitar al número de líneas solicitadas
                all_lines = all_lines[-lines:]
        
        return {"logs": "".join(all_lines[-lines:])}
    except Exception as e:
        return {"logs": f"Error al leer logs: {str(e)}"}

# Endpoint para controlar el servicio
@app.post("/api/cucm/service/{action}")
async def control_cucm_service(action: str, user=Depends(admin_only)):
    if isinstance(user, RedirectResponse):
        return user
        
    if action not in ["start", "stop", "restart"]:
        raise HTTPException(status_code=400, detail=f"Acción no válida: {action}")
    
    try:
        # Verificar si systemctl está disponible
        if not os.path.exists("/bin/systemctl"):
            raise HTTPException(status_code=500, detail="El control de servicios systemd no está disponible en este sistema")
        
        # Ejecutar el comando correspondiente
        result = os.system(f"systemctl {action} tarificador-cucm.service")
        
        if result != 0:
            raise HTTPException(status_code=500, detail=f"Error al {action} el servicio")
        
        # Actualizar el estado en la base de datos
        db = SessionLocal()
        
        status_map = {
            "start": "starting",
            "stop": "stopping",
            "restart": "restarting"
        }
        
        update_query = text("""
            UPDATE cucm_config SET 
            last_status = :status,
            last_status_update = CURRENT_TIMESTAMP
            WHERE id = (SELECT id FROM cucm_config ORDER BY id DESC LIMIT 1)
        """)
        
        db.execute(update_query, {"status": status_map[action]})
        db.commit()
        db.close()
        
        action_name = {"start": "iniciar", "stop": "detener", "restart": "reiniciar"}[action]
        return {"message": f"Orden para {action_name} el servicio enviada correctamente"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al controlar el servicio: {str(e)}")
    
# Endpoint para obtener configuración del servicio
@app.get("/api/config")
def get_service_config():
    # Retorna la configuración del servicio
    return {
        "cucm.host": "10.224.0.10",
        "cucm.user": "jtapiuser",
        "cucm.password": "fr4v4t3l",
        "cucm.appinfo": "TarificadorApp",
        "api.url": "http://localhost:8000",
        "monitor.extensions": "all",
        "reconnect.delay": "60",
        "log.level": "INFO"
    }

@app.post("/cdr/rejected")
def rejected_call(event: CallEvent):
    db = SessionLocal()
    
    # Registrar la llamada rechazada
    cdr = CDR(
        calling_number=event.calling_number,
        called_number=event.called_number,
        start_time=event.start_time,
        end_time=event.end_time,
        duration_seconds=0,
        cost=0,
        status="rejected_insufficient_balance"
    )
    db.add(cdr)
    db.commit()
    db.close()
    
    return {"message": "Llamada rechazada registrada"}

@app.get("/dashboard/monitoreo")
async def monitoreo(request: Request):
    # Verifica si el usuario está autenticado
    user = await authenticated_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=302)
    
    db = SessionLocal()
    
    try:
        # Estadísticas básicas para mostrar inicialmente
        today = datetime.now().date()

        llamadas_hoy = db.execute(
            text("SELECT COUNT(*) FROM cdr WHERE DATE(start_time) = '{today}'")
        ).scalar() or 0
        
        minutos_hoy = db.execute(
            text("SELECT COALESCE(SUM(duration_seconds)/60, 0) FROM cdr WHERE DATE(start_time) = '{today}'")
        ).scalar() or 0
        
        alertas_saldo = db.execute(
            text("SELECT calling_number, saldo FROM saldo_anexos WHERE saldo < 5.0 ORDER BY saldo ASC LIMIT 5")
        ).fetchall()
        
        llamadas_recientes = db.execute(
            text("""
            SELECT * FROM cdr 
            ORDER BY start_time DESC 
            LIMIT 10
            """)
        ).fetchall()
    
    finally:
        db.close()
    
    return templates.TemplateResponse("monitoreo.html", {
        "request": request,
        "user": user,
        "llamadas_hoy": llamadas_hoy,
        "minutos_hoy": round(float(minutos_hoy), 1),
        "alertas_saldo": alertas_saldo,
        "llamadas_recientes": llamadas_recientes
    })

# Nuevos endpoints para el módulo de zonas, prefijos y tarifas
@app.get("/dashboard/zonas")
async def dashboard_zonas(request: Request, user=Depends(admin_only)):
    if isinstance(user, RedirectResponse):
        return user
        
    db = SessionLocal()
    
    query = text("SELECT id, nombre, descripcion FROM zonas ORDER BY nombre")
    zonas = db.execute(query).fetchall()
    
    db.close()
    
    return templates.TemplateResponse("dashboard_zonas.html", {
        "request": request, "zonas": zonas, "user": user
    })

@app.get("/dashboard/prefijos")
async def dashboard_prefijos(request: Request, zona_id: int = None, user=Depends(admin_only)):
    if isinstance(user, RedirectResponse):
        return user
        
    db = SessionLocal()
    
    zonas_query = text("SELECT id, nombre FROM zonas ORDER BY nombre")
    zonas = db.execute(zonas_query).fetchall()
    
    if zona_id:
        prefijos_query = text("""
            SELECT p.id, p.zona_id, p.prefijo, p.longitud_minima, p.longitud_maxima, z.nombre as zona_nombre
            FROM prefijos p
            JOIN zonas z ON p.zona_id = z.id
            WHERE p.zona_id = :zona_id
            ORDER BY p.prefijo
        """)
        prefijos = db.execute(prefijos_query, {"zona_id": zona_id}).fetchall()
        
        zona_query = text("SELECT id, nombre FROM zonas WHERE id = :zona_id")
        zona_actual = db.execute(zona_query, {"zona_id": zona_id}).fetchone()
    else:
        prefijos_query = text("""
            SELECT p.id, p.zona_id, p.prefijo, p.longitud_minima, p.longitud_maxima, z.nombre as zona_nombre
            FROM prefijos p
            JOIN zonas z ON p.zona_id = z.id
            ORDER BY z.nombre, p.prefijo
        """)
        prefijos = db.execute(prefijos_query).fetchall()
        zona_actual = None
    
    db.close()
    
    return templates.TemplateResponse("dashboard_prefijos.html", {
        "request": request, "prefijos": prefijos, "zonas": zonas, 
        "zona_actual": zona_actual, "user": user
    })

@app.get("/dashboard/tarifas")
async def dashboard_tarifas(request: Request, zona_id: int = None, user=Depends(admin_only)):
    if isinstance(user, RedirectResponse):
        return user
        
    db = SessionLocal()
    
    zonas_query = text("SELECT id, nombre FROM zonas ORDER BY nombre")
    zonas = db.execute(zonas_query).fetchall()
    
    if zona_id:
        tarifas_query = text("""
            SELECT t.id, t.zona_id, t.tarifa_segundo, t.fecha_inicio, t.activa, z.nombre as zona_nombre
            FROM tarifas t
            JOIN zonas z ON t.zona_id = z.id
            WHERE t.zona_id = :zona_id
            ORDER BY t.fecha_inicio DESC
        """)
        tarifas = db.execute(tarifas_query, {"zona_id": zona_id}).fetchall()
        
        zona_query = text("SELECT id, nombre FROM zonas WHERE id = :zona_id")
        zona_actual = db.execute(zona_query, {"zona_id": zona_id}).fetchone()
    else:
        tarifas_query = text("""
            SELECT t.id, t.zona_id, t.tarifa_segundo, t.fecha_inicio, t.activa, z.nombre as zona_nombre
            FROM tarifas t
            JOIN zonas z ON t.zona_id = z.id
            ORDER BY z.nombre, t.fecha_inicio DESC
        """)
        tarifas = db.execute(tarifas_query).fetchall()
        zona_actual = None
    
    db.close()
    
    return templates.TemplateResponse("dashboard_tarifas.html", {
        "request": request, "tarifas": tarifas, "zonas": zonas, 
        "zona_actual": zona_actual, "user": user
    })

@app.get("/dashboard/estadisticas_zona")
def dashboard_estadisticas_zona(request: Request, 
                               user=Depends(authenticated_user),
                               fecha_inicio: str = Query(None),
                               fecha_fin: str = Query(None)):
    db = SessionLocal()
    # Establecer fechas por defecto si no se proporcionan
    if not fecha_inicio:
        fecha_inicio = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
    if not fecha_fin:
        fecha_fin = datetime.now().strftime('%Y-%m-%d')
    
    # Consulta SQL
    query = f"""
        SELECT 
            z.nombre as zona_nombre,
            COUNT(*) as total_llamadas,
            SUM(c.duration_seconds) / 60.0 as duracion_total_minutos,
            SUM(c.cost) as costo_total,
            AVG(c.cost) as costo_promedio,
            AVG(c.duration_seconds) / 60.0 as duracion_promedio_minutos
        FROM cdr c
        JOIN prefijos p ON SUBSTR(c.called_number, 1, LENGTH(p.prefijo)) = p.prefijo
        JOIN zonas z ON p.zona_id = z.id
        WHERE c.start_time >= '{fecha_inicio} 00:00:00'
          AND c.start_time <= '{fecha_fin} 23:59:59'
        GROUP BY z.id, z.nombre
        ORDER BY costo_total DESC
    """
    estadisticas = db.execute(text(query)).fetchall()
    db.close()

    # Calcular totales y porcentajes
    total_costo = sum(float(e.costo_total) if e.costo_total else 0 for e in estadisticas)

    # Crear nuevas estadísticas con el campo porcentaje_total
    estadisticas_modificadas = []
    for stat in estadisticas:
        costo_total = float(stat.costo_total) if stat.costo_total else 0
        porcentaje_total = (costo_total / total_costo * 100) if total_costo > 0 else 0

        estadisticas_modificadas.append({
            "zona_nombre": stat.zona_nombre,
            "total_llamadas": stat.total_llamadas,
            "duracion_total_minutos": stat.duracion_total_minutos,
            "costo_total": costo_total,
            "costo_promedio": stat.costo_promedio,
            "duracion_promedio_minutos": stat.duracion_promedio_minutos,
            "porcentaje_total": porcentaje_total
        })

    # Preparar datos para gráficos
    zonas_labels = [e.zona_nombre for e in estadisticas] if estadisticas else []
    llamadas_data = [e.total_llamadas for e in estadisticas] if estadisticas else []
    costo_data = [float(e.costo_total) for e in estadisticas] if estadisticas else []

    return templates.TemplateResponse("dashboard_estadisticas_zona.html", {
        "request": request,
        "user": user,
        "estadisticas": estadisticas_modificadas,
        "total_llamadas": sum(e.total_llamadas for e in estadisticas) if estadisticas else 0,
        "total_minutos": sum(e.duracion_total_minutos for e in estadisticas) if estadisticas else 0,
        "total_costo": total_costo,
        "zonas_activas": len(estadisticas),
        "zonas_labels": zonas_labels,
        "llamadas_data": llamadas_data,
        "costo_data": costo_data,
        "fecha_inicio": fecha_inicio,
        "fecha_fin": fecha_fin
    })

@app.get("/export/cdr/pdf")
def export_cdr_pdf(
    user=Depends(admin_only),
    start_date: str = Query(None),
    end_date: str = Query(None),
    calling_number: str = Query(None),
    called_number: str = Query(None)
):
    """Exporta los registros CDR a un archivo PDF."""
    from sqlalchemy import text
    from datetime import datetime
    from jinja2 import Template
    from weasyprint import HTML
    import io
    
    db = SessionLocal()
    
    try:
        # Construir la consulta base
        query = """
            SELECT c.calling_number, c.called_number, 
                   c.start_time, c.end_time, c.duration_seconds, 
                   c.duration_billable, c.cost, c.status,
                   z.nombre as zona
            FROM cdr c
            LEFT JOIN zonas z ON c.zona_id = z.id
            WHERE 1=1
        """
        
        params = {}
        
        # Agregar filtros si se proporcionan
        if calling_number:
            query += " AND c.calling_number = :calling_number"
            params["calling_number"] = calling_number
        
        if called_number:
            query += " AND c.called_number = :called_number"
            params["called_number"] = called_number
        
        if start_date:
            query += " AND c.start_time >= :start_date"
            params["start_date"] = f"{start_date} 00:00:00"
        
        if end_date:
            query += " AND c.end_time <= :end_date"
            params["end_date"] = f"{end_date} 23:59:59"
        
        # Ordenar por fecha descendente
        query += " ORDER BY c.start_time DESC LIMIT 1000"
        
        # Ejecutar la consulta usando text()
        rows = db.execute(text(query), params).fetchall()
        
        # Crear el HTML para el reporte
        html_template = """
        <html>
        <head>
            <style>
                body { font-family: Arial, sans-serif; }
                table { width: 100%; border-collapse: collapse; }
                th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
                th { background-color: #f2f2f2; }
                .header { text-align: center; margin-bottom: 20px; }
                .footer { text-align: center; margin-top: 20px; font-size: 0.8em; }
            </style>
        </head>
        <body>
            <div class="header">
                <h1>Reporte de Llamadas (CDR)</h1>
                <p>Fecha: {{ datetime.now().strftime('%Y-%m-%d %H:%M:%S') }}</p>
                {% if start_date or end_date %}
                <p>Período: {{ start_date or 'Inicio' }} - {{ end_date or 'Fin' }}</p>
                {% endif %}
            </div>
            
            <table>
                <thead>
                    <tr>
                        <th>Origen</th>
                        <th>Destino</th>
                        <th>Fecha/Hora</th>
                        <th>Duración</th>
                        <th>Facturada</th>
                        <th>Zona</th>
                        <th>Costo</th>
                    </tr>
                </thead>
                <tbody>
                {% for row in rows %}
                    <tr>
                        <td>{{ row[0] }}</td>
                        <td>{{ row[1] }}</td>
                        <td>{{ row[2].strftime('%Y-%m-%d %H:%M:%S') }}</td>
                        <td>{{ "%d:%02d"|format(row[4]//60, row[4]%60) }}</td>
                        <td>{{ "%d:%02d"|format(row[5]//60, row[5]%60) }}</td>
                        <td>{{ row[8] or 'N/A' }}</td>
                        <td>S/{{ "%.4f"|format(row[6]) }}</td>
                    </tr>
                {% endfor %}
                </tbody>
            </table>
            
            <div class="footer">
                <p>Reporte generado el {{ datetime.now().strftime('%Y-%m-%d %H:%M:%S') }}</p>
            </div>
        </body>
        </html>
        """
        
        # Renderizar el HTML con Jinja2
        template = Template(html_template)
        html_content = template.render(
            rows=rows, 
            datetime=datetime,
            start_date=start_date,
            end_date=end_date
        )
        
        # Generar el PDF usando WeasyPrint
        pdf = HTML(string=html_content).write_pdf()
        
        filename = f"cdr_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        db.close()
        
        # Devolver el PDF como respuesta
        return StreamingResponse(
            io.BytesIO(pdf),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        db.close()
        import traceback
        traceback.print_exc()
        return {"error": str(e)}

@app.get("/export/cdr/csv")
def export_cdr_csv(
    user=Depends(admin_only),
    start_date: str = Query(None),
    end_date: str = Query(None),
    calling_number: str = Query(None),
    called_number: str = Query(None)
):
    """Exporta los registros CDR a un archivo CSV."""
    from sqlalchemy import text
    from datetime import datetime
    import csv
    import io
    
    db = SessionLocal()
    
    try:
        # Construir la consulta base
        query = """
            SELECT c.calling_number, c.called_number, 
                   c.start_time, c.end_time, c.duration_seconds, 
                   c.duration_billable, c.cost, c.status,
                   z.nombre as zona_nombre, c.connect_time,
                   c.dialing_time, c.network_reached_time, c.network_alerting_time,
                   c.direction, c.release_cause
            FROM cdr c
            LEFT JOIN zonas z ON c.zona_id = z.id
            WHERE 1=1
        """
        
        params = {}
        
        # Agregar filtros si se proporcionan
        if calling_number:
            query += " AND c.calling_number = :calling_number"
            params["calling_number"] = calling_number
        
        if called_number:
            query += " AND c.called_number = :called_number"
            params["called_number"] = called_number
        
        if start_date:
            query += " AND c.start_time >= :start_date"
            params["start_date"] = f"{start_date} 00:00:00"
        
        if end_date:
            query += " AND c.end_time <= :end_date"
            params["end_date"] = f"{end_date} 23:59:59"
        
        # Ordenar por fecha descendente
        query += " ORDER BY c.start_time DESC LIMIT 10000"
        
        # Ejecutar la consulta usando text()
        rows = db.execute(text(query), params).fetchall()
        
        # Crear el archivo CSV en memoria
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Escribir encabezados
        #se retiro columna row[7]: Estado
        writer.writerow([
            'Origen', 'Destino', 'Fecha/Hora Inicio', 'Fecha/Hora Fin', 
            'Duracion Total (seg)', 'Duracion Facturable (seg)', 'Costo', 
            'Zona', 'Hora Contestado', 'Hora Marcacion',
            'Hora Alcance Red', 'Hora Timbrando', 'Direccion', 'Codigo Liberacion'
        ])
        
        # Escribir datos
        for row in rows:
            # Formatear fechas y valores numéricos
            start_time = row[2].strftime('%Y-%m-%d %H:%M:%S') if row[2] else ''
            end_time = row[3].strftime('%Y-%m-%d %H:%M:%S') if row[3] else ''
            connect_time = row[9].strftime('%Y-%m-%d %H:%M:%S') if row[9] else ''
            dialing_time = row[10].strftime('%Y-%m-%d %H:%M:%S') if row[10] else ''
            network_reached_time = row[11].strftime('%Y-%m-%d %H:%M:%S') if row[11] else ''
            network_alerting_time = row[12].strftime('%Y-%m-%d %H:%M:%S') if row[12] else ''
            
            #se retiro columna row[7]: status
            writer.writerow([
                row[0],                     # calling_number
                row[1],                     # called_number
                start_time,                 # start_time
                end_time,                   # end_time
                row[4],                     # duration_seconds
                row[5],                     # duration_billable
                f"{float(row[6]):.4f}",     # cost
                row[8],                     # zona_nombre
                connect_time,               # connect_time
                dialing_time,               # dialing_time
                network_reached_time,       # network_reached_time
                network_alerting_time,      # network_alerting_time
                row[13],                    # direction
                row[14]                     # release_cause
            ])
        
        db.close()
        
        # Configurar el nombre del archivo
        filename = f"cdr_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        # Devolver el CSV como respuesta
        output.seek(0)
        return StreamingResponse(
            io.StringIO(output.getvalue()),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        db.close()
        import traceback
        traceback.print_exc()
        return {"error": str(e)}

@app.get("/export/cdr/excel")
def export_cdr_excel(
    user=Depends(admin_only),
    start_date: str = Query(None),
    end_date: str = Query(None),
    calling_number: str = Query(None),
    called_number: str = Query(None)
):
    """Exporta los registros CDR a un archivo Excel (XLS)."""
    from sqlalchemy import text
    from datetime import datetime
    import io
    
    db = SessionLocal()
    
    try:
        # Construir la consulta base
        query = """
            SELECT c.calling_number, c.called_number, 
                   c.start_time, c.end_time, c.duration_seconds, 
                   c.duration_billable, c.cost, c.status,
                   z.nombre as zona_nombre, c.connect_time,
                   c.dialing_time, c.network_reached_time, c.network_alerting_time,
                   c.direction, c.release_cause
            FROM cdr c
            LEFT JOIN zonas z ON c.zona_id = z.id
            WHERE 1=1
        """
        
        params = {}
        
        # Agregar filtros si se proporcionan
        if calling_number:
            query += " AND c.calling_number = :calling_number"
            params["calling_number"] = calling_number
        
        if called_number:
            query += " AND c.called_number = :called_number"
            params["called_number"] = called_number
        
        if start_date:
            query += " AND c.start_time >= :start_date"
            params["start_date"] = f"{start_date} 00:00:00"
        
        if end_date:
            query += " AND c.end_time <= :end_date"
            params["end_date"] = f"{end_date} 23:59:59"
        
        # Ordenar por fecha descendente
        query += " ORDER BY c.start_time DESC LIMIT 10000"
        
        # Ejecutar la consulta usando text()
        rows = db.execute(text(query), params).fetchall()
        
        # Lista de encabezados
        headers = [
            'Origen', 'Destino', 'Fecha/Hora Inicio', 'Fecha/Hora Fin', 
            'Duración Total (seg)', 'Duración Facturable (seg)', 'Costo', 
            'Zona', 'Hora Contestado', 'Hora Marcación',
            'Hora Alcance Red', 'Hora Timbrando', 'Dirección', 'Código Liberación'
        ]
        
        # Preparar los datos para Excel
        excel_data = []
        
        # Añadir encabezados
        excel_data.append(headers)
        
        # Añadir filas de datos
        for row in rows:
            # Formatear fechas y valores numéricos
            start_time = row[2].strftime('%Y-%m-%d %H:%M:%S') if row[2] else ''
            end_time = row[3].strftime('%Y-%m-%d %H:%M:%S') if row[3] else ''
            connect_time = row[9].strftime('%Y-%m-%d %H:%M:%S') if row[9] else ''
            dialing_time = row[10].strftime('%Y-%m-%d %H:%M:%S') if row[10] else ''
            network_reached_time = row[11].strftime('%Y-%m-%d %H:%M:%S') if row[11] else ''
            network_alerting_time = row[12].strftime('%Y-%m-%d %H:%M:%S') if row[12] else ''
            
            excel_data.append([
                row[0],                     # calling_number
                row[1],                     # called_number
                start_time,                 # start_time
                end_time,                   # end_time
                row[4],                     # duration_seconds
                row[5],                     # duration_billable
                float(row[6]),              # cost (como número para Excel)
                row[8],                     # zona_nombre
                connect_time,               # connect_time
                dialing_time,               # dialing_time
                network_reached_time,       # network_reached_time
                network_alerting_time,      # network_alerting_time
                row[13],                    # direction
                row[14]                     # release_cause
            ])
        
        # Nombre del archivo
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"cdr_report_{timestamp}.xls"
        
        # Intentar usar diferentes bibliotecas para generar el Excel
        excel_binary = None
        
        # Intento 1: Usar xlwt (mejor para XLS)
        try:
            import xlwt
            
            # Crear libro y hoja
            workbook = xlwt.Workbook()
            worksheet = workbook.add_sheet('CDR Report')
            
            # Estilos
            header_style = xlwt.easyxf('font: bold on; align: wrap on, vert centre, horiz center; pattern: pattern solid, fore_color gray25')
            date_style = xlwt.easyxf(num_format_str='YYYY-MM-DD HH:MM:SS')
            number_style = xlwt.easyxf(num_format_str='0.0000')
            
            # Establecer anchos de columna
            for i in range(len(headers)):
                worksheet.col(i).width = 256 * 20  # Aproximadamente 20 caracteres de ancho
            
            # Escribir datos
            for row_idx, row_data in enumerate(excel_data):
                for col_idx, cell_value in enumerate(row_data):
                    # Aplicar estilos según el tipo de datos
                    if row_idx == 0:  # Encabezados
                        worksheet.write(row_idx, col_idx, cell_value, header_style)
                    elif col_idx in [2, 3, 8, 9, 10, 11]:  # Columnas de fecha
                        worksheet.write(row_idx, col_idx, cell_value, date_style)
                    elif col_idx == 6:  # Columna de costo
                        worksheet.write(row_idx, col_idx, cell_value, number_style)
                    else:
                        worksheet.write(row_idx, col_idx, cell_value)
            
            # Guardar a un BytesIO
            output = io.BytesIO()
            workbook.save(output)
            excel_binary = output.getvalue()
            
            print("✅ Excel generado con xlwt")
            
        except ImportError:
            # xlwt no está disponible, intentar con openpyxl
            try:
                import openpyxl
                from openpyxl.styles import Font, Alignment, PatternFill
                from openpyxl.utils import get_column_letter
                
                # Crear libro y hoja
                workbook = openpyxl.Workbook()
                worksheet = workbook.active
                worksheet.title = 'CDR Report'
                
                # Estilos
                header_font = Font(bold=True)
                header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                header_fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
                
                # Establecer anchos de columna
                for i in range(len(headers)):
                    worksheet.column_dimensions[get_column_letter(i+1)].width = 20
                
                # Escribir datos
                for row_idx, row_data in enumerate(excel_data, 1):  # openpyxl es 1-based
                    for col_idx, cell_value in enumerate(row_data, 1):
                        cell = worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
                        
                        # Aplicar estilos a los encabezados
                        if row_idx == 1:
                            cell.font = header_font
                            cell.alignment = header_alignment
                            cell.fill = header_fill
                
                # Guardar a un BytesIO
                output = io.BytesIO()
                workbook.save(output)
                excel_binary = output.getvalue()
                
                print("✅ Excel generado con openpyxl")
                
            except ImportError:
                # openpyxl no está disponible, intentar con pandas
                try:
                    import pandas as pd
                    
                    # Convertir datos a DataFrame (omitir la fila de encabezados)
                    df = pd.DataFrame(excel_data[1:], columns=headers)
                    
                    # Guardar a un BytesIO
                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                        df.to_excel(writer, sheet_name='CDR Report', index=False)
                        
                        # Dar formato a la hoja
                        workbook = writer.book
                        worksheet = writer.sheets['CDR Report']
                        
                        # Formato para encabezados
                        header_format = workbook.add_format({
                            'bold': True,
                            'bg_color': '#DDDDDD',
                            'align': 'center',
                            'valign': 'vcenter',
                            'text_wrap': True
                        })
                        
                        # Aplicar formato a encabezados
                        for col_num, value in enumerate(df.columns.values):
                            worksheet.write(0, col_num, value, header_format)
                            worksheet.set_column(col_num, col_num, 20)
                    
                    excel_binary = output.getvalue()
                    
                    print("✅ Excel generado con pandas")
                    
                except ImportError:
                    # Si ninguna biblioteca Excel está disponible, generar CSV como alternativa
                    print("⚠️ No hay bibliotecas Excel disponibles, generando CSV")
                    import csv
                    
                    output = io.StringIO()
                    writer = csv.writer(output)
                    
                    # Escribir todas las filas
                    for row_data in excel_data:
                        writer.writerow(row_data)
                    
                    # Cambiar el nombre de archivo a CSV
                    filename = f"cdr_report_{timestamp}.csv"
                    
                    # Devolver el CSV como respuesta
                    output.seek(0)
                    db.close()
                    return StreamingResponse(
                        io.StringIO(output.getvalue()),
                        media_type="text/csv",
                        headers={"Content-Disposition": f"attachment; filename={filename}"}
                    )
        
        db.close()
        
        # Devolver el Excel como respuesta
        return StreamingResponse(
            io.BytesIO(excel_binary),
            media_type="application/vnd.ms-excel",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        db.close()
        import traceback
        traceback.print_exc()
        return {"error": str(e)}
                
# Exportar reporte de consumo por zona
@app.get("/export/consumo_zona/pdf")
async def export_consumo_zona_pdf(user=Depends(admin_only)):
    if isinstance(user, RedirectResponse):
        return user
    
    db = SessionLocal()
    
    # Estadísticas de consumo por zona (últimos 30 días)
    stats_query = text("""
        SELECT z.nombre, COUNT(c.id) as total_llamadas, 
               SUM(c.duration_seconds) as total_duracion, 
               SUM(c.cost) as total_costo
        FROM cdr c
        JOIN zonas z ON c.zona_id = z.id
        WHERE c.start_time >= NOW() - INTERVAL '30 days'
        GROUP BY z.nombre
        ORDER BY total_costo DESC
    """)
    estadisticas = db.execute(stats_query).fetchall()
    db.close()

    html_template = """
    <html>
    <body>
    <h1>Reporte de Consumo por Zona</h1>
    <h3>Últimos 30 días</h3>
    <table border="1">
        <thead>
            <tr>
                <th>Zona</th>
                <th>Total Llamadas</th>
                <th>Duración Total (seg)</th>
                <th>Costo Total</th>
            </tr>
        </thead>
        <tbody>
        {% for row in estadisticas %}
            <tr>
                <td>{{ row[0] }}</td>
                <td>{{ row[1] }}</td>
                <td>{{ row[2] if row[2] else 0 }}</td>
                <td>${{ "%.2f"|format(row[3] if row[3] else 0) }}</td>
            </tr>
        {% endfor %}
        </tbody>
    </table>
    </body>
    </html>
    """
    template = Template(html_template)
    html_content = template.render(estadisticas=estadisticas)

    pdf = HTML(string=html_content).write_pdf()

    return StreamingResponse(
        io.BytesIO(pdf),
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=consumo_zona_report.pdf"}
    )

# API para el módulo de zonas
@app.get("/api/zonas")
async def listar_zonas(user=Depends(authenticated_user)):
    if isinstance(user, RedirectResponse):
        return user
        
    db = SessionLocal()
    query = text("SELECT id, nombre, descripcion FROM zonas ORDER BY nombre")
    zonas = db.execute(query).fetchall()
    
    result = []
    for zona in zonas:
        result.append({
            "id": zona[0],
            "nombre": zona[1],
            "descripcion": zona[2]
        })
    
    db.close()
    return result

@app.post("/api/zonas")
async def crear_zona(zona: ZonaCreate, user=Depends(admin_only)):
    if isinstance(user, RedirectResponse):
        return user
        
    db = SessionLocal()
    
    # Verificar que no exista otra zona con el mismo nombre
    check_query = text("SELECT id FROM zonas WHERE nombre = :nombre")
    existing = db.execute(check_query, {"nombre": zona.nombre}).fetchone()
    
    if existing:
        db.close()
        raise HTTPException(status_code=400, detail="Ya existe una zona con ese nombre")
    
    # Crear la zona
    insert_query = text("""
        INSERT INTO zonas (nombre, descripcion)
        VALUES (:nombre, :descripcion)
        RETURNING id
    """)
    
    result = db.execute(insert_query, {
        "nombre": zona.nombre,
        "descripcion": zona.descripcion
    })
    
    zona_id = result.fetchone()[0]
    
    # Crear una tarifa por defecto para la zona
    insert_tarifa_query = text("""
        INSERT INTO tarifas (zona_id, tarifa_segundo, fecha_inicio, activa)
        VALUES (:zona_id, :tarifa_segundo, CURRENT_TIMESTAMP, TRUE)
    """)
    
    db.execute(insert_tarifa_query, {
        "zona_id": zona_id,
        "tarifa_segundo": 0.0005  # Tarifa por defecto: 0.03 por minuto
    })
    
    db.commit()
    db.close()
    
    return {"id": zona_id, "nombre": zona.nombre, "descripcion": zona.descripcion}

@app.put("/api/zonas/{zona_id}")
async def actualizar_zona(zona_id: int, zona: ZonaCreate, user=Depends(admin_only)):
    if isinstance(user, RedirectResponse):
        return user
        
    db = SessionLocal()
    
    # Verificar que la zona exista
    check_query = text("SELECT id FROM zonas WHERE id = :zona_id")
    existing = db.execute(check_query, {"zona_id": zona_id}).fetchone()
    
    if not existing:
        db.close()
        raise HTTPException(status_code=404, detail="Zona no encontrada")
    
    # Verificar que no exista otra zona con el mismo nombre
    check_name_query = text("SELECT id FROM zonas WHERE nombre = :nombre AND id != :zona_id")
    existing_name = db.execute(check_name_query, {"nombre": zona.nombre, "zona_id": zona_id}).fetchone()
    
    if existing_name:
        db.close()
        raise HTTPException(status_code=400, detail="Ya existe otra zona con ese nombre")
    
    # Actualizar la zona
    update_query = text("""
        UPDATE zonas
        SET nombre = :nombre, descripcion = :descripcion
        WHERE id = :zona_id
    """)
    
    db.execute(update_query, {
        "zona_id": zona_id,
        "nombre": zona.nombre,
        "descripcion": zona.descripcion
    })
    
    db.commit()
    db.close()
    
    return {"id": zona_id, "nombre": zona.nombre, "descripcion": zona.descripcion}

@app.delete("/api/zonas/{zona_id}")
async def eliminar_zona(zona_id: int, user=Depends(admin_only)):
    if isinstance(user, RedirectResponse):
        return user
        
    db = SessionLocal()
    
    # Verificar que la zona exista
    check_query = text("SELECT id FROM zonas WHERE id = :zona_id")
    existing = db.execute(check_query, {"zona_id": zona_id}).fetchone()
    
    if not existing:
        db.close()
        raise HTTPException(status_code=404, detail="Zona no encontrada")
    
    # Verificar si tiene prefijos o tarifas asociadas
    check_prefijos_query = text("SELECT COUNT(*) FROM prefijos WHERE zona_id = :zona_id")
    prefijos_count = db.execute(check_prefijos_query, {"zona_id": zona_id}).scalar()
    
    check_tarifas_query = text("SELECT COUNT(*) FROM tarifas WHERE zona_id = :zona_id")
    tarifas_count = db.execute(check_tarifas_query, {"zona_id": zona_id}).scalar()
    
    if prefijos_count > 0 or tarifas_count > 0:
        db.close()
        raise HTTPException(
            status_code=400, 
            detail="No se puede eliminar la zona porque tiene prefijos o tarifas asociadas"
        )
    
    # Eliminar la zona
    delete_query = text("DELETE FROM zonas WHERE id = :zona_id")
    db.execute(delete_query, {"zona_id": zona_id})
    
    db.commit()
    db.close()
    
    return {"message": "Zona eliminada correctamente"}

# API para el módulo de prefijos
@app.get("/api/prefijos")
async def listar_prefijos(zona_id: int = None, user=Depends(authenticated_user)):
    if isinstance(user, RedirectResponse):
        return user
        
    db = SessionLocal()
    
    if zona_id:
        query = text("""
            SELECT p.id, p.zona_id, p.prefijo, p.longitud_minima, p.longitud_maxima, z.nombre as zona_nombre
            FROM prefijos p
            JOIN zonas z ON p.zona_id = z.id
            WHERE p.zona_id = :zona_id
            ORDER BY p.prefijo
        """)
        prefijos = db.execute(query, {"zona_id": zona_id}).fetchall()
    else:
        query = text("""
            SELECT p.id, p.zona_id, p.prefijo, p.longitud_minima, p.longitud_maxima, z.nombre as zona_nombre
            FROM prefijos p
            JOIN zonas z ON p.zona_id = z.id
            ORDER BY z.nombre, p.prefijo
        """)
        prefijos = db.execute(query).fetchall()
    
    result = []
    for prefijo in prefijos:
        result.append({
            "id": prefijo[0],
            "zona_id": prefijo[1],
            "prefijo": prefijo[2],
            "longitud_minima": prefijo[3],
            "longitud_maxima": prefijo[4],
            "zona_nombre": prefijo[5]
        })
    
    db.close()
    return result

@app.post("/api/prefijos")
async def crear_prefijo(prefijo: PrefijoCreate, user=Depends(admin_only)):
    if isinstance(user, RedirectResponse):
        return user
        
    db = SessionLocal()
    
    # Verificar que la zona exista
    check_zona_query = text("SELECT id FROM zonas WHERE id = :zona_id")
    existing_zona = db.execute(check_zona_query, {"zona_id": prefijo.zona_id}).fetchone()
    
    if not existing_zona:
        db.close()
        raise HTTPException(status_code=404, detail="Zona no encontrada")
    
    # Validar longitudes
    if prefijo.longitud_minima > prefijo.longitud_maxima:
        db.close()
        raise HTTPException(status_code=400, detail="La longitud mínima no puede ser mayor que la longitud máxima")
    
    # Insertar el prefijo
    insert_query = text("""
        INSERT INTO prefijos (zona_id, prefijo, longitud_minima, longitud_maxima)
        VALUES (:zona_id, :prefijo, :longitud_minima, :longitud_maxima)
        RETURNING id
    """)
    
    result = db.execute(insert_query, {
        "zona_id": prefijo.zona_id,
        "prefijo": prefijo.prefijo,
        "longitud_minima": prefijo.longitud_minima,
        "longitud_maxima": prefijo.longitud_maxima
    })
    
    prefijo_id = result.fetchone()[0]
    
    # Obtener el nombre de la zona
    zona_query = text("SELECT nombre FROM zonas WHERE id = :zona_id")
    zona_nombre = db.execute(zona_query, {"zona_id": prefijo.zona_id}).fetchone()[0]
    
    db.commit()
    db.close()
    
    return {
        "id": prefijo_id,
        "zona_id": prefijo.zona_id,
        "prefijo": prefijo.prefijo,
        "longitud_minima": prefijo.longitud_minima,
        "longitud_maxima": prefijo.longitud_maxima,
        "zona_nombre": zona_nombre
    }

@app.put("/api/prefijos/{prefijo_id}")
async def actualizar_prefijo(prefijo_id: int, prefijo: PrefijoCreate, user=Depends(admin_only)):
    if isinstance(user, RedirectResponse):
        return user
        
    db = SessionLocal()
    
    # Verificar que el prefijo exista
    check_query = text("SELECT id FROM prefijos WHERE id = :prefijo_id")
    existing = db.execute(check_query, {"prefijo_id": prefijo_id}).fetchone()
    
    if not existing:
        db.close()
        raise HTTPException(status_code=404, detail="Prefijo no encontrado")
    
    # Verificar que la zona exista
    check_zona_query = text("SELECT id FROM zonas WHERE id = :zona_id")
    existing_zona = db.execute(check_zona_query, {"zona_id": prefijo.zona_id}).fetchone()
    
    if not existing_zona:
        db.close()
        raise HTTPException(status_code=404, detail="Zona no encontrada")
    
    # Validar longitudes
    if prefijo.longitud_minima > prefijo.longitud_maxima:
        db.close()
        raise HTTPException(status_code=400, detail="La longitud mínima no puede ser mayor que la longitud máxima")
    
    # Actualizar el prefijo
    update_query = text("""
        UPDATE prefijos
        SET zona_id = :zona_id, prefijo = :prefijo, longitud_minima = :longitud_minima, longitud_maxima = :longitud_maxima
        WHERE id = :prefijo_id
    """)
    
    db.execute(update_query, {
        "prefijo_id": prefijo_id,
        "zona_id": prefijo.zona_id,
        "prefijo": prefijo.prefijo,
        "longitud_minima": prefijo.longitud_minima,
        "longitud_maxima": prefijo.longitud_maxima
    })
    
    # Obtener el nombre de la zona
    zona_query = text("SELECT nombre FROM zonas WHERE id = :zona_id")
    zona_nombre = db.execute(zona_query, {"zona_id": prefijo.zona_id}).fetchone()[0]
    
    db.commit()
    db.close()
    
    return {
        "id": prefijo_id,
        "zona_id": prefijo.zona_id,
        "prefijo": prefijo.prefijo,
        "longitud_minima": prefijo.longitud_minima,
        "longitud_maxima": prefijo.longitud_maxima,
        "zona_nombre": zona_nombre
    }

@app.delete("/api/prefijos/{prefijo_id}")
async def eliminar_prefijo(prefijo_id: int, user=Depends(admin_only)):
    if isinstance(user, RedirectResponse):
        return user
        
    db = SessionLocal()
    
    # Verificar que el prefijo exista
    check_query = text("SELECT id FROM prefijos WHERE id = :prefijo_id")
    existing = db.execute(check_query, {"prefijo_id": prefijo_id}).fetchone()
    
    if not existing:
        db.close()
        raise HTTPException(status_code=404, detail="Prefijo no encontrado")
    
    # Eliminar el prefijo
    delete_query = text("DELETE FROM prefijos WHERE id = :prefijo_id")
    db.execute(delete_query, {"prefijo_id": prefijo_id})
    
    db.commit()
    db.close()
    
    return {"message": "Prefijo eliminado correctamente"}

# API para el módulo de tarifas
@app.get("/api/tarifas")
async def listar_tarifas(zona_id: int = None, user=Depends(authenticated_user)):
    if isinstance(user, RedirectResponse):
        return user
        
    db = SessionLocal()
    
    if zona_id:
        query = text("""
            SELECT t.id, t.zona_id, t.tarifa_segundo, t.fecha_inicio, t.activa, z.nombre as zona_nombre
            FROM tarifas t
            JOIN zonas z ON t.zona_id = z.id
            WHERE t.zona_id = :zona_id
            ORDER BY t.fecha_inicio DESC
        """)
        tarifas = db.execute(query, {"zona_id": zona_id}).fetchall()
    else:
        query = text("""
            SELECT t.id, t.zona_id, t.tarifa_segundo, t.fecha_inicio, t.activa, z.nombre as zona_nombre
            FROM tarifas t
            JOIN zonas z ON t.zona_id = z.id
            ORDER BY z.nombre, t.fecha_inicio DESC
        """)
        tarifas = db.execute(query).fetchall()
    
    result = []
    for tarifa in tarifas:
        result.append({
            "id": tarifa[0],
            "zona_id": tarifa[1],
            "tarifa_segundo": float(tarifa[2]),
            "fecha_inicio": tarifa[3].isoformat() if tarifa[3] else None,
            "activa": tarifa[4],
            "zona_nombre": tarifa[5]
        })
    
    db.close()
    return result

@app.post("/api/tarifas")
async def crear_tarifa(tarifa: TarifaCreate, user=Depends(admin_only)):
    if isinstance(user, RedirectResponse):
        return user
        
    db = SessionLocal()
    
    # Verificar que la zona exista
    check_zona_query = text("SELECT id FROM zonas WHERE id = :zona_id")
    existing_zona = db.execute(check_zona_query, {"zona_id": tarifa.zona_id}).fetchone()
    
    if not existing_zona:
        db.close()
        raise HTTPException(status_code=404, detail="Zona no encontrada")
    
    # Desactivar las tarifas anteriores de esta zona
    update_query = text("""
        UPDATE tarifas 
        SET activa = FALSE 
        WHERE zona_id = :zona_id AND activa = TRUE
    """)
    
    db.execute(update_query, {"zona_id": tarifa.zona_id})
    
    # Insertar la nueva tarifa
    insert_query = text("""
        INSERT INTO tarifas (zona_id, tarifa_segundo, fecha_inicio, activa)
        VALUES (:zona_id, :tarifa_segundo, CURRENT_TIMESTAMP, TRUE)
        RETURNING id, fecha_inicio
    """)
    
    result = db.execute(insert_query, {
        "zona_id": tarifa.zona_id,
        "tarifa_segundo": tarifa.tarifa_segundo
    })
    
    id_fecha = result.fetchone()
    tarifa_id = id_fecha[0]
    fecha_inicio = id_fecha[1]
    
    # Obtener el nombre de la zona
    zona_query = text("SELECT nombre FROM zonas WHERE id = :zona_id")
    zona_nombre = db.execute(zona_query, {"zona_id": tarifa.zona_id}).fetchone()[0]
    
    db.commit()
    db.close()
    
    return {
        "id": tarifa_id,
        "zona_id": tarifa.zona_id,
        "tarifa_segundo": tarifa.tarifa_segundo,
        "fecha_inicio": fecha_inicio.isoformat() if fecha_inicio else None,
        "activa": True,
        "zona_nombre": zona_nombre
    }

@app.put("/api/tarifas/{tarifa_id}/activar")
async def activar_tarifa(tarifa_id: int, user=Depends(admin_only)):
    if isinstance(user, RedirectResponse):
        return user
        
    db = SessionLocal()
    
    # Verificar que la tarifa exista
    check_query = text("SELECT id, zona_id FROM tarifas WHERE id = :tarifa_id")
    existing = db.execute(check_query, {"tarifa_id": tarifa_id}).fetchone()
    
    if not existing:
        db.close()
        raise HTTPException(status_code=404, detail="Tarifa no encontrada")
    
    zona_id = existing[1]
    
    # Desactivar las tarifas anteriores de esta zona
    update_other_query = text("""
        UPDATE tarifas 
        SET activa = FALSE 
        WHERE zona_id = :zona_id AND id != :tarifa_id AND activa = TRUE
    """)
    
    db.execute(update_other_query, {"zona_id": zona_id, "tarifa_id": tarifa_id})
    
    # Activar esta tarifa
    update_query = text("""
        UPDATE tarifas 
        SET activa = TRUE 
        WHERE id = :tarifa_id
    """)
    
    db.execute(update_query, {"tarifa_id": tarifa_id})
    
    # Obtener datos actualizados de la tarifa
    tarifa_query = text("""
        SELECT t.id, t.zona_id, t.tarifa_segundo, t.fecha_inicio, t.activa, z.nombre as zona_nombre
        FROM tarifas t
        JOIN zonas z ON t.zona_id = z.id
        WHERE t.id = :tarifa_id
    """)
    
    tarifa = db.execute(tarifa_query, {"tarifa_id": tarifa_id}).fetchone()
    
    db.commit()
    db.close()
    
    return {
        "id": tarifa[0],
        "zona_id": tarifa[1],
        "tarifa_segundo": float(tarifa[2]),
        "fecha_inicio": tarifa[3].isoformat() if tarifa[3] else None,
        "activa": tarifa[4],
        "zona_nombre": tarifa[5],
        "message": "Tarifa activada correctamente"
    }

@app.delete("/api/tarifas/{tarifa_id}")
async def eliminar_tarifa(tarifa_id: int, user=Depends(admin_only)):
    if isinstance(user, RedirectResponse):
        return user
        
    db = SessionLocal()
    
    # Verificar que la tarifa exista
    check_query = text("SELECT id, zona_id, activa FROM tarifas WHERE id = :tarifa_id")
    existing = db.execute(check_query, {"tarifa_id": tarifa_id}).fetchone()
    
    if not existing:
        db.close()
        raise HTTPException(status_code=404, detail="Tarifa no encontrada")
    
    zona_id = existing[1]
    is_active = existing[2]
    
    # Si la tarifa está activa, verificar que haya otra tarifa que se pueda activar
    if is_active:
        check_others_query = text("""
            SELECT COUNT(*) 
            FROM tarifas 
            WHERE zona_id = :zona_id AND id != :tarifa_id
        """)
        
        other_count = db.execute(check_others_query, {"zona_id": zona_id, "tarifa_id": tarifa_id}).scalar()
        
        if other_count == 0:
            db.close()
            raise HTTPException(
                status_code=400, 
                detail="No se puede eliminar la única tarifa de la zona"
            )
        
        # Activar la tarifa más reciente de la zona
        activate_query = text("""
            UPDATE tarifas 
            SET activa = TRUE 
            WHERE zona_id = :zona_id AND id != :tarifa_id 
            ORDER BY fecha_inicio DESC 
            LIMIT 1
        """)
        
        db.execute(activate_query, {"zona_id": zona_id, "tarifa_id": tarifa_id})
    
    # Eliminar la tarifa
    delete_query = text("DELETE FROM tarifas WHERE id = :tarifa_id")
    db.execute(delete_query, {"tarifa_id": tarifa_id})
    
    db.commit()
    db.close()
    
    return {"message": "Tarifa eliminada correctamente"}


class FacCode(Base):
    __tablename__ = "fac_codes"
    id = Column(Integer, primary_key=True, index=True)
    authorization_code = Column(String, unique=True, index=True)
    authorization_code_name = Column(String)  # Nombre descriptivo para CUCM
    authorization_level = Column(Integer)     # Nivel de autorización (0-255)
    description = Column(String, nullable=True)
    active = Column(Boolean, default=True)
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    cucm_synced = Column(Boolean, default=False)  # Indicador de sincronización
    
class FacAudit(Base):
    __tablename__ = "fac_audit"
    id = Column(Integer, primary_key=True, index=True)
    authorization_code = Column(String)
    action = Column(String)  # 'create', 'update', 'delete', 'sync', 'sync_create', 'sync_update'
    admin_user = Column(String)  # Usuario que realizó la acción
    timestamp = Column(DateTime, default=datetime.utcnow)
    details = Column(String, nullable=True)
    success = Column(Boolean, default=True)


# Modelos de entrada/salida para códigos FAC
class FacCodeBase(BaseModel):
    authorization_code: str = Field(..., min_length=1, max_length=16)
    authorization_code_name: str = Field(..., min_length=1, max_length=50)
    authorization_level: int = Field(..., ge=0, le=255)
    description: Optional[str] = None
    active: bool = True

class FacCodeCreate(FacCodeBase):
    pass

class FacCodeUpdate(BaseModel):
    authorization_code_name: Optional[str] = Field(None, min_length=1, max_length=50)
    authorization_level: Optional[int] = Field(None, ge=0, le=255)
    description: Optional[str] = None
    active: Optional[bool] = None

class FacCodeResponse(FacCodeBase):
    id: int
    created_at: datetime
    updated_at: datetime
    cucm_synced: bool

    class Config:
        orm_mode = True

# Modelos para auditoría
class FacAuditBase(BaseModel):
    authorization_code: str
    action: str
    admin_user: str
    details: Optional[str] = None
    success: bool = True

class FacAuditCreate(FacAuditBase):
    pass

class FacAuditResponse(FacAuditBase):
    id: int
    timestamp: datetime

    class Config:
        orm_mode = True

# Modelo para resultados de sincronización
class SyncResult(BaseModel):
    message: str
    status: str
    created: Optional[int] = None
    updated: Optional[int] = None
    errors: Optional[int] = None

# Dependency para obtener la sesión DB
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

class UserAuthCode(Base):
    __tablename__ = "user_auth_codes"
    id = Column(Integer, primary_key=True, index=True)
    extension = Column(String, unique=True, index=True)
    auth_code = Column(String, unique=True)
    auth_level = Column(Integer)
    description = Column(String, nullable=True)
    active = Column(Boolean, default=True)
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class UserAuthCodeAudit(Base):
    __tablename__ = "user_auth_code_audit"
    id = Column(Integer, primary_key=True, index=True)
    extension = Column(String)
    auth_code = Column(String)
    action = Column(String)  # 'create', 'update', 'delete'
    admin_user = Column(String)  # Usuario administrador que realizó la acción
    timestamp = Column(DateTime, default=datetime.utcnow)
    details = Column(String, nullable=True)

class UserAuthCodeBase(BaseModel):
    extension: str
    auth_code: str
    auth_level: int
    description: Optional[str] = None
    active: bool = True

class UserAuthCodeCreate(UserAuthCodeBase):
    pass

class UserAuthCodeUpdate(BaseModel):
    auth_code: Optional[str] = None
    auth_level: Optional[int] = None
    description: Optional[str] = None
    active: Optional[bool] = None

class UserAuthCodeResponse(UserAuthCodeBase):
    id: int
    created_at: datetime
    updated_at: datetime

    class Config:
        orm_mode = True

from zeep import Client, Settings
from zeep.transports import Transport
from requests import Session
from requests.auth import HTTPBasicAuth
import urllib3
from zeep.exceptions import Fault

# Desactivar advertencias SSL
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Función para conectar con CUCM
def get_cucm_client():
    """Crea un cliente SOAP para conectarse a CUCM"""
    # Cargar configuración
    CUCM_ADDRESS = os.getenv('CUCM_ADDRESS', '10.224.0.10')
    CUCM_USERNAME = os.getenv('CUCM_USERNAME', 'admin')
    CUCM_PASSWORD = os.getenv('CUCM_PASSWORD', 'fr4v4t3l')
    WSDL_FILE = 'schema/AXLAPI.wsdl'
    
    # Configurar sesión
    session = Session()
    session.verify = False
    session.auth = HTTPBasicAuth(CUCM_USERNAME, CUCM_PASSWORD)
    
    # Configurar transporte
    transport = Transport(session=session, timeout=10)
    settings = Settings(strict=False, xml_huge_tree=True)
    
    # Crear cliente
    client = Client(WSDL_FILE, settings=settings, transport=transport)
    
    # Crear servicio
    service = client.create_service(
        '{http://www.cisco.com/AXLAPIService/}AXLAPIBinding',
        f'https://{CUCM_ADDRESS}:8443/axl/'
    )
    
    return service

# Clase para gestionar códigos FAC
class CucmFacManager:
    """Gestor de códigos de autorización forzada (FAC) para CUCM"""
    
    def __init__(self):
        self.client = get_cucm_client()

    def list_fac_info(self):
        """
        Lista todos los códigos de autorización forzada
        Usando la variante 2 que funciona correctamente
        """
        try:
            print("Llamando a listFacInfo con returnedTags y searchCriteria como keywords")
            response = self.client.listFacInfo(
                searchCriteria={"name": "%"},
                returnedTags={"name": "", "code": "", "authorizationLevel": ""}
            )
            
            # Debug detallado para inspeccionar la respuesta
            print(f"Tipo de respuesta: {type(response)}")
            print(f"Contenido de respuesta: {response}")
            
            # Inspeccionar atributos de la respuesta
            if hasattr(response, '__dict__'):
                print(f"Atributos de respuesta: {dir(response)}")
            
            return response
        except Exception as e:
            print(f"Error en list_fac_info: {e}")
            import traceback
            print(traceback.format_exc())
            return None

    def process_fac_info(response):
        """
        Procesa la respuesta de CUCM y devuelve una lista de códigos FAC
        Adaptado a la estructura específica de la respuesta
        """
        fac_list = []
        print(f"Procesando respuesta de tipo: {type(response)}")
        
        if not response:
            print("Respuesta vacía")
            return fac_list
        
        # Usar getattr para acceder al atributo 'return' que es una palabra reservada
        response_return = getattr(response, 'return', None)
        if not response_return:
            print("No se encontró el atributo 'return' en la respuesta")
            return fac_list
        
        # Acceder a facInfo dentro del atributo return
        fac_info = getattr(response_return, 'facInfo', None)
        if not fac_info:
            print("No se encontró 'facInfo' en la respuesta")
            return fac_list
        
        # Procesar la lista de códigos FAC
        if isinstance(fac_info, list):
            print(f"facInfo es una lista con {len(fac_info)} elementos")
            for fac in fac_info:
                fac_data = {
                    'uuid': fac.get('uuid', ''),
                    'name': fac.get('name', ''),
                    'code': fac.get('code', ''),
                    'level': fac.get('authorizationLevel', 0),
                    'source': 'cucm'
                }
                print(f"Procesado FAC: {fac_data}")
                fac_list.append(fac_data)
        else:
            # Procesar un único elemento
            print("facInfo es un único elemento")
            fac_data = {
                'uuid': getattr(fac_info, 'uuid', ''),
                'name': getattr(fac_info, 'name', ''),
                'code': getattr(fac_info, 'code', ''),
                'level': getattr(fac_info, 'authorizationLevel', 0),
                'source': 'cucm'
            }
            print(f"Procesado FAC: {fac_data}")
            fac_list.append(fac_data)
        
        print(f"Total de FACs procesados: {len(fac_list)}")
        return fac_list

    def add_fac_info(self, code, name, auth_level):
        """
        Añade un nuevo código de autorización forzada
        Adaptado exactamente del ejemplo PHP proporcionado
        """
        try:
            # Crear la estructura exacta del ejemplo PHP
            fac_info = {
                "name": name,
                "code": code,
                "authorizationLevel": auth_level
            }
            
            # Llamada igual al ejemplo PHP: $client->addFacInfo(array("facInfo"=>$facInfo))
            response = self.client.addFacInfo(facInfo=fac_info)
            return True
        except Fault as e:
            print(f"SOAP Fault al añadir FAC: {e}")
            return False
        except Exception as e:
            print(f"Error general al añadir FAC: {e}")
            return False
    
    def update_fac_info(self, code, name=None, auth_level=None):
        """Actualiza un código de autorización existente"""
        try:
            # Crear el objeto para la actualización
            update_data = {}
            
            # Siempre incluir el código (requerido para identificar el FAC)
            update_data["code"] = code
            
            if name is not None:
                update_data["name"] = name
            
            if auth_level is not None:
                update_data["authorizationLevel"] = auth_level
            
            # Estructura similar a addFacInfo
            response = self.client.updateFacInfo(facInfo=update_data)
            return True
        except Exception as e:
            print(f"Error al actualizar FAC {code}: {e}")
            return False
    
    def remove_fac_info(self, code):
        """Elimina un código de autorización forzada"""
        try:
            # Estructura según el patrón establecido
            response = self.client.removeFacInfo(code=code)
            return True
        except Exception as e:
            print(f"Error al eliminar FAC {code}: {e}")
            return False
        

# IMPLEMENTACIÓN DE FAC (FORCED AUTHORIZATION CODES)
@app.get("/dashboard/fac")
async def dashboard_fac(request: Request, user=Depends(admin_only)):
    """Dashboard para gestionar códigos de autorización forzada (FAC)"""
    if isinstance(user, RedirectResponse):
        return user
    
    db = SessionLocal()
    # Obtener códigos FAC de la base de datos local
    local_fac_list = []
    try:
        local_codes = db.query(FacCode).all()
        for code in local_codes:
            local_fac_list.append({
                'id': code.id,
                'code': code.authorization_code,
                'name': code.authorization_code_name,
                'level': code.authorization_level,
                'description': code.description,
                'active': code.active,
                'cucm_synced': code.cucm_synced,
                'source': 'local'  # Indicar que proviene de la BD local
            })
    except Exception as e:
        print(f"Error obteniendo códigos FAC locales: {e}")
    
    # Intentar obtener códigos del CUCM también
    cucm_fac_list = []
    try:
        # Instanciar el gestor
        fac_manager = CucmFacManager()
        
        # Obtener lista de códigos FAC de CUCM
        response = fac_manager.list_fac_info()

        if response:
            try:
                # Procesar resultados
                cucm_fac_list = process_fac_info(response)
            except Exception as e:
                print(f"Error procesando resultados FAC desde CUCM: {e}")
    except Exception as e:
        print(f"Error conectando con CUCM: {e}")
    
    # Cerrar la conexión a la base de datos
    db.close()
    
    # Combinar ambas listas (preferimos los datos locales si hay duplicados)
    combined_list = local_fac_list.copy()
    
    # Agregar códigos de CUCM que no están en la lista local
    local_codes = {item['code'] for item in local_fac_list}
    for cucm_item in cucm_fac_list:
        if cucm_item['code'] not in local_codes:
            combined_list.append(cucm_item)
    
    return templates.TemplateResponse("dashboard_fac.html", {
        "request": request, 
        "fac_list": combined_list, 
        "user": user,
        "local_count": len(local_fac_list),
        "cucm_count": len(cucm_fac_list),
        "total_count": len(combined_list)
    })

@app.get("/api/fac")
async def get_fac_list(user=Depends(admin_only)):
    """API para obtener todos los códigos FAC"""
    if isinstance(user, RedirectResponse):
        return user
    
    fac_manager = CucmFacManager()
    response = fac_manager.list_fac_info()
    
    # Procesar resultados
    fac_list = process_fac_info(response)

    return {"fac_codes": fac_list}

def process_fac_info(response):
    """
    Procesa la respuesta de CUCM y devuelve una lista de códigos FAC
    """
    fac_list = []

    if not response or not hasattr(response, 'return'):
        return fac_list

    result_data = getattr(response, 'return', None)

    if not hasattr(result_data, 'facInfo'):
        return fac_list

    fac_info = result_data.facInfo

    # Si hay múltiples resultados, es una lista
    if isinstance(fac_info, list):
        for fac in fac_info:
            fac_list.append({
                "uuid": getattr(fac, 'uuid', None),
                "name": getattr(fac, 'name', None),
                "code": getattr(fac, 'code', None),
                "level": getattr(fac, 'authorizationLevel', None)
            })
    else:
        # Solo un resultado
        fac_list.append({
            "uuid": getattr(fac_info, 'uuid', None),
            "name": getattr(fac_info, 'name', None),
            "code": getattr(fac_info, 'code', None),
            "level": getattr(fac_info, 'authorizationLevel', None)
        })

    return fac_list

@app.post("/api/fac")
async def create_fac(
    fac: FacCodeCreate,
    user: dict = Depends(admin_only)  # Asegúrate de que esto devuelva un usuario autenticado
):
    """API para crear un nuevo código FAC"""
    #if isinstance(user, RedirectResponse):
    #    return user
    
    # Instanciar el gestor
    fac_manager = CucmFacManager()
    
    # Registrar en la base de datos local
    db = SessionLocal()
    try:
        # Verificar si ya existe
        existing = db.query(FacCode).filter(FacCode.authorization_code == fac.authorization_code).first()
        if existing:
            db.close()
            raise HTTPException(status_code=400, detail="El código de autorización ya existe")
        
        # Crear nuevo código FAC en la BD local
        db_fac = FacCode(
            authorization_code=fac.authorization_code,
            authorization_code_name=fac.authorization_code_name,
            authorization_level=fac.authorization_level,
            description="",
            active=True,
            cucm_synced=False
        )
        db.add(db_fac)
        
        # Registrar en auditoría
        audit = FacAudit(
            authorization_code=fac.authorization_code,
            action="create",
            admin_user=user["username"],
            details="Creación manual desde dashboard",
            success=True
        )
        db.add(audit)
        db.commit()
        
        # Intentar crear en CUCM
        success = fac_manager.add_fac_info(fac.authorization_code, fac.authorization_code_name, fac.authorization_level)
        
        if success:
            # Actualizar estado de sincronización
            db_fac.cucm_synced = True
            
            # Registrar en auditoría
            audit = FacAudit(
                authorization_code=fac.authorization_code,
                action="sync_create",
                admin_user=user["username"],
                details="Creado automáticamente en CUCM",
                success=True
            )
            db.add(audit)
            db.commit()
        
        db.close()
        return {"success": True, "message": "Código FAC creado exitosamente", "synced": success}
    except HTTPException:
        db.close()
        raise
    except Exception as e:
        db.rollback()
        db.close()
        raise HTTPException(status_code=500, detail=f"Error al crear código FAC: {str(e)}")

@app.put("/api/fac/{code}")
async def update_fac(
    code: str,
    name: Optional[str] = Form(None),
    auth_level: Optional[int] = Form(None),
    description: Optional[str] = Form(None),
    active: Optional[bool] = Form(None),
    user=Depends(admin_only)
):
    """API para actualizar un código FAC existente"""
    if isinstance(user, RedirectResponse):
        return user
    
    db = SessionLocal()
    try:
        # Buscar el código en la BD local
        fac = db.query(FacCode).filter(FacCode.authorization_code == code).first()
        if not fac:
            db.close()
            raise HTTPException(status_code=404, detail="Código FAC no encontrado")
        
        # Actualizar campos proporcionados
        if name is not None:
            fac.authorization_code_name = name
        
        if auth_level is not None:
            fac.authorization_level = auth_level
            
        if description is not None:
            fac.description = description
            
        if active is not None:
            fac.active = active
        
        # Marcar como no sincronizado
        fac.cucm_synced = False
        
        # Registrar en auditoría
        audit = FacAudit(
            authorization_code=code,
            action="update",
            admin_user=user["username"],
            details="Actualización manual desde dashboard",
            success=True
        )
        db.add(audit)
        db.commit()
        
        # Intentar actualizar en CUCM si está activo
        if fac.active:
            fac_manager = CucmFacManager()
            success = fac_manager.update_fac_info(code, name, auth_level)
            
            if success:
                # Actualizar estado de sincronización
                fac.cucm_synced = True
                
                # Registrar en auditoría
                audit = FacAudit(
                    authorization_code=code,
                    action="sync_update",
                    admin_user=user["username"],
                    details="Actualizado automáticamente en CUCM",
                    success=True
                )
                db.add(audit)
                db.commit()
        
        db.close()
        return {"success": True, "message": f"Código FAC {code} actualizado exitosamente"}
    except HTTPException:
        db.close()
        raise
    except Exception as e:
        db.rollback()
        db.close()
        raise HTTPException(status_code=500, detail=f"Error al actualizar código FAC: {str(e)}")

@app.delete("/api/fac/{code}")
async def delete_fac(code: str, user=Depends(admin_only)):
    """API para eliminar un código FAC"""
    if isinstance(user, RedirectResponse):
        return user
    
    db = SessionLocal()
    try:
        # Buscar el código en la BD local
        fac = db.query(FacCode).filter(FacCode.authorization_code == code).first()
        if not fac:
            db.close()
            raise HTTPException(status_code=404, detail="Código FAC no encontrado")
        
        # Eliminar de CUCM primero
        fac_manager = CucmFacManager()
        cucm_success = fac_manager.remove_fac_info(code)
        
        # Eliminar de la BD local
        db.delete(fac)
        
        # Registrar en auditoría
        audit = FacAudit(
            authorization_code=code,
            action="delete",
            admin_user=user["username"],
            details=f"Eliminación manual desde dashboard (CUCM: {'éxito' if cucm_success else 'fallido'})",
            success=True
        )
        db.add(audit)
        db.commit()
        
        db.close()
        return {"success": True, "message": f"Código FAC {code} eliminado exitosamente"}
    except HTTPException:
        db.close()
        raise
    except Exception as e:
        db.rollback()
        db.close()
        raise HTTPException(status_code=500, detail=f"Error al eliminar código FAC: {str(e)}")
    
@app.get("/api/fac/test-connection")
async def test_fac_connection(user=Depends(admin_only)):
    """API para probar la conexión con CUCM y listar códigos FAC"""
    if isinstance(user, RedirectResponse):
        return user
    
    try:
        # Probar la conexión utilizando el mismo patrón que el ejemplo PHP
        client = get_cucm_client()
        
        # Construir la solicitud exactamente como en PHP
        returned_tags = {
            "name": "",
            "code": "",
            "authorizationLevel": ""
        }
        
        search_criteria = {
            "name": "%"  # Wildcard para encontrar todos los FAC
        }
        
        try:
            # Llamada equivalente a la del ejemplo PHP
            response = client.listFacInfo(
                returnedTags=returned_tags,
                searchCriteria=search_criteria
            )
            
            # Analizar la estructura de la respuesta para depuración
            response_info = {
                "type": type(response).__name__,
                "has_return_": hasattr(response, "return_")
            }
            
            if hasattr(response, "return_"):
                return_obj = response.return_
                response_info["return_type"] = type(return_obj).__name__
                response_info["has_facInfo"] = hasattr(return_obj, "facInfo")
                
                if hasattr(return_obj, "facInfo"):
                    fac_info = return_obj.facInfo
                    response_info["facInfo_type"] = type(fac_info).__name__
                    response_info["facInfo_is_list"] = isinstance(fac_info, list)
                    
                    if isinstance(fac_info, list):
                        response_info["facInfo_count"] = len(fac_info)
                        if len(fac_info) > 0:
                            sample = fac_info[0]
                            response_info["sample_fac"] = {
                                "has_uuid": hasattr(sample, "uuid"),
                                "has_name": hasattr(sample, "name"),
                                "has_code": hasattr(sample, "code"),
                                "has_authLevel": hasattr(sample, "authorizationLevel")
                            }
                    else:
                        response_info["sample_fac"] = {
                            "has_uuid": hasattr(fac_info, "uuid"),
                            "has_name": hasattr(fac_info, "name"),
                            "has_code": hasattr(fac_info, "code"),
                            "has_authLevel": hasattr(fac_info, "authorizationLevel")
                        }
            
            return {
                "success": True, 
                "message": "Conexión exitosa con CUCM",
                "response_info": response_info
            }
            
        except Fault as e:
            # Error SOAP específico
            return {
                "success": False,
                "message": f"Error de SOAP: {str(e)}",
                "error_type": "soap_fault"
            }
            
    except Exception as e:
        # Error general de conexión
        return {
            "success": False,
            "message": f"Error de conexión: {str(e)}",
            "error_type": "connection_error"
        }


@app.get("/api/fac/raw-test")
async def test_fac_raw(user=Depends(admin_only)):
    """Prueba directa con SOAP crudo para diagnosticar el problema"""
    if isinstance(user, RedirectResponse):
        return user
    
    try:
        # Configuración
        cucm_address = "10.224.0.10"
        cucm_username = "admin" 
        cucm_password = "fr4v4t3l"
        
        # URL del servicio AXL
        url = f"https://{cucm_address}:8443/axl/"
        
        # Encabezados SOAP
        headers = {
            "Content-Type": "text/xml; charset=utf-8",
            "SOAPAction": "CUCM:DB ver=12.5 listFacInfo"
        }
        
        # Cuerpo SOAP con la estructura exacta basada en el ejemplo PHP
        body = """
        <soapenv:Envelope xmlns:soapenv="http://schemas.xmlsoap.org/soap/envelope/" 
                         xmlns:ns="http://www.cisco.com/AXL/API/12.5">
           <soapenv:Header/>
           <soapenv:Body>
              <ns:listFacInfo>
                 <searchCriteria>
                    <name>%</name>
                 </searchCriteria>
                 <returnedTags>
                    <name/>
                    <code/>
                    <authorizationLevel/>
                 </returnedTags>
              </ns:listFacInfo>
           </soapenv:Body>
        </soapenv:Envelope>
        """
        
        # Realizar la solicitud
        session = Session()
        session.verify = False
        session.auth = HTTPBasicAuth(cucm_username, cucm_password)
        
        response = session.post(url, headers=headers, data=body)
        
        # Imprimir respuesta
        response_text = response.text
        
        return {
            "status_code": response.status_code,
            "response": response_text
        }
    except Exception as e:
        return {"error": str(e)}


# Sincronización con CUCM
@app.post("/api/fac/sync-with-cucm")
async def sync_fac_with_cucm(
    background_tasks: BackgroundTasks, 
    db = Depends(lambda: SessionLocal()), 
    user=Depends(admin_only)
):
    """Inicia la sincronización de FAC con CUCM en segundo plano"""
    if isinstance(user, RedirectResponse):
        return user
        
    try:
        # Obtener nombre de usuario
        username = user["username"] if isinstance(user, dict) and "username" in user else "sistema"
        
        # Iniciar tarea en segundo plano
        background_tasks.add_task(sync_all_fac_with_cucm, username, db)
        
        return {"message": "Sincronización iniciada en segundo plano", "status": "success"}
    except Exception as e:
        import traceback
        print(f"Error al iniciar sincronización: {str(e)}")
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Error al iniciar sincronización: {str(e)}")

# Función de sincronización que se ejecuta en segundo plano
def sync_all_fac_with_cucm(admin_username: str, db: SessionLocal):
    """Sincroniza todos los códigos FAC con CUCM"""
    try:
        # Obtener códigos FAC desde la base de datos
        fac_codes = db.query(FacCode).filter(FacCode.active == True).all()
        
        # Conectar con CUCM
        fac_manager = CucmFacManager()
        
        # Obtener códigos existentes en CUCM
        cucm_codes = {}
        try:
            # Listar códigos FAC existentes en CUCM
            response = fac_manager.listFacInfo()
            
            # Procesar respuesta para obtener códigos existentes
            if hasattr(response, 'return_') and hasattr(response.return_, 'facInfo'):
                fac_info = response.return_.facInfo
                
                # Si es una lista
                if isinstance(fac_info, list):
                    for fac in fac_info:
                        if hasattr(fac, 'code') and hasattr(fac, 'name'):
                            cucm_codes[fac.code] = {
                                'name': fac.name,
                                'level': fac.authorizationLevel if hasattr(fac, 'authorizationLevel') else None
                            }
                # Si es un solo objeto
                elif hasattr(fac_info, 'code') and hasattr(fac_info, 'name'):
                    cucm_codes[fac_info.code] = {
                        'name': fac_info.name,
                        'level': fac_info.authorizationLevel if hasattr(fac_info, 'authorizationLevel') else None
                    }
        except Exception as e:
            # Registrar error al obtener códigos existentes
            audit_entry = FacAudit(
                authorization_code="N/A",
                action="sync",
                admin_user=admin_username,
                details=f"Error obteniendo códigos de CUCM: {str(e)}",
                success=False
            )
            db.add(audit_entry)
            db.commit()
            print(f"Error obteniendo códigos de CUCM: {str(e)}")
        
        # Sincronizar códigos
        created_count = 0
        updated_count = 0
        error_count = 0
        
        for code in fac_codes:
            try:
                if code.authorization_code in cucm_codes:
                    # Verificar si necesita actualización
                    cucm_code = cucm_codes[code.authorization_code]
                    needs_update = (
                        code.authorization_code_name != cucm_code['name'] or
                        code.authorization_level != cucm_code['level']
                    )
                    
                    if needs_update:
                        # Actualizar código existente usando el método corregido
                        success = fac_manager.update_fac_info(
                            code.authorization_code,
                            code.authorization_code_name,
                            code.authorization_level
                        )
                        
                        if success:
                            # Registrar auditoría
                            audit_entry = FacAudit(
                                authorization_code=code.authorization_code,
                                action="sync_update",
                                admin_user=admin_username,
                                details=f"Código actualizado en CUCM",
                                success=True
                            )
                            db.add(audit_entry)
                            updated_count += 1
                            
                            # Marcar como sincronizado
                            code.cucm_synced = True
                            db.commit()
                        else:
                            raise Exception("Fallo al actualizar en CUCM")
                else:
                    # Crear nuevo código usando el método corregido
                    success = fac_manager.add_fac_info(
                        code.authorization_code,
                        code.authorization_code_name,
                        code.authorization_level
                    )
                    
                    if success:
                        # Registrar auditoría
                        audit_entry = FacAudit(
                            authorization_code=code.authorization_code,
                            action="sync_create",
                            admin_user=admin_username,
                            details=f"Código creado en CUCM",
                            success=True
                        )
                        db.add(audit_entry)
                        created_count += 1
                        
                        # Marcar como sincronizado
                        code.cucm_synced = True
                        db.commit()
                    else:
                        raise Exception("Fallo al crear en CUCM")
                
            except Exception as e:
                # Registrar error
                error_count += 1
                audit_entry = FacAudit(
                    authorization_code=code.authorization_code,
                    action="sync",
                    admin_user=admin_username,
                    details=f"Error de sincronización: {str(e)}",
                    success=False
                )
                db.add(audit_entry)
                db.commit()
                print(f"Error sincronizando código {code.authorization_code}: {str(e)}")
                
        # Verificar si hay códigos en CUCM que no están en nuestro sistema
        orphan_count = 0
        for cucm_code in cucm_codes:
            if not db.query(FacCode).filter(FacCode.authorization_code == cucm_code).first():
                orphan_count += 1
                audit_entry = FacAudit(
                    authorization_code=cucm_code,
                    action="sync_detect",
                    admin_user=admin_username,
                    details=f"Código encontrado en CUCM pero no en sistema local",
                    success=True
                )
                db.add(audit_entry)
                db.commit()
        
        # Registrar resumen final
        summary_entry = FacAudit(
            authorization_code="SUMMARY",
            action="sync_complete",
            admin_user=admin_username,
            details=f"Sincronización completada. Creados: {created_count}, Actualizados: {updated_count}, Errores: {error_count}, Huérfanos detectados: {orphan_count}",
            success=True
        )
        db.add(summary_entry)
        db.commit()
        print(f"Sincronización completada. Creados: {created_count}, Actualizados: {updated_count}, Errores: {error_count}, Huérfanos: {orphan_count}")
                
    except Exception as e:
        # Registrar error general
        audit_entry = FacAudit(
            authorization_code="N/A",
            action="sync",
            admin_user=admin_username,
            details=f"Error general de sincronización: {str(e)}",
            success=False
        )
        db.add(audit_entry)
        db.commit()
        print(f"Error general en sincronización: {str(e)}")


@app.get("/dashboard/fac/historial")
async def fac_historial(
    request: Request, 
    limit: int = Query(100, ge=10, le=500),
    user=Depends(admin_only)
):
    """Muestra el historial de cambios y sincronización de FAC"""
    if isinstance(user, RedirectResponse):
        return user
        
    db = SessionLocal()
    
    # Obtener registros de auditoría
    try:
        # Consultar registros de auditoría ordenados por fecha descendente
        audit_logs = db.query(FacAudit).order_by(FacAudit.timestamp.desc()).limit(limit).all()
        
        # Preparar estadísticas resumidas
        stats = {
            "total_records": len(audit_logs),
            "success_count": sum(1 for log in audit_logs if log.success),
            "error_count": sum(1 for log in audit_logs if not log.success),
            "by_action": {}
        }
        
        # Contar registros por tipo de acción
        action_counts = {}
        for log in audit_logs:
            action = log.action
            if action not in action_counts:
                action_counts[action] = 0
            action_counts[action] += 1
        
        stats["by_action"] = action_counts
        
    except Exception as e:
        print(f"Error al consultar registros de auditoría: {e}")
        audit_logs = []
        stats = {
            "error": str(e),
            "total_records": 0
        }
    
    db.close()
    
    return templates.TemplateResponse("dashboard_fac_historial.html", {
        "request": request,
        "audit_logs": audit_logs,
        "stats": stats,
        "user": user,
        "limit": limit
    })

@app.get("/dashboard/fac/sync")
async def dashboard_fac_sync(request: Request, user=Depends(admin_only)):
    """Dashboard para sincronización de FAC con CUCM"""
    if isinstance(user, RedirectResponse):
        return user
        
    db = SessionLocal()
    
    # Obtener estadísticas de sincronización
    stats = {
        "total_local": db.query(FacCode).count(),
        "total_synced": db.query(FacCode).filter(FacCode.cucm_synced == True).count()
    }
    
    # Obtener últimos eventos de auditoría
    audit_logs = db.query(FacAudit).order_by(FacAudit.timestamp.desc()).limit(20).all()
    
    # Obtener última sincronización completa
    last_sync = db.query(FacAudit).filter(
        FacAudit.action == "sync_complete"
    ).order_by(FacAudit.timestamp.desc()).first()
    
    db.close()
    
    return templates.TemplateResponse("dashboard_fac_sync.html", {
        "request": request,
        "stats": stats,
        "audit_logs": audit_logs,
        "last_sync": last_sync,
        "user": user
    })

@app.post("/api/fac/sync-from-cucm")
async def sync_fac_from_cucm(background_tasks: BackgroundTasks, user=Depends(admin_only)):
    """Importar códigos FAC desde CUCM a la base de datos local"""
    if isinstance(user, RedirectResponse):
        return user
        
    background_tasks.add_task(import_fac_from_cucm, user["username"])
    
    return {"message": "Importación iniciada en segundo plano", "status": "success"}

def import_fac_from_cucm(admin_username: str):
    """Importa códigos FAC desde CUCM a la base de datos local"""
    db = SessionLocal()
    
    try:
        # Conectar con CUCM
        client = get_cucm_client()

        # Listar códigos existentes en CUCM
        returned_tags = {
            "name": "",
            "code": "",
            "authorizationLevel": ""
        }
        
        search_criteria = {
            "name": "%"  # Wildcard para encontrar todos los FAC
        }
        
        # Listar códigos FAC existentes en CUCM
        response = client.listFacInfo(
            returnedTags=returned_tags,
            searchCriteria=search_criteria
        )
        
        # Procesar respuesta para obtener códigos existentes
        created_count = 0
        updated_count = 0
        unchanged_count = 0
        
        if hasattr(response, 'return_') and hasattr(response.return_, 'facInfo'):
            fac_info = response.return_.facInfo
            
            # Procesar lista de códigos
            fac_list = []
            if isinstance(fac_info, list):
                fac_list = fac_info
            else:
                fac_list = [fac_info]
            
            # Importar cada código
            for fac in fac_list:
                if hasattr(fac, 'code') and hasattr(fac, 'name'):
                    code = fac.code
                    name = fac.name
                    level = fac.authorizationLevel if hasattr(fac, 'authorizationLevel') else 0
                    
                    # Verificar si el código ya existe
                    existing = db.query(FacCode).filter(FacCode.authorization_code == code).first()
                    
                    if existing:
                        # Verificar si necesita actualización
                        if (existing.authorization_code_name != name or 
                            existing.authorization_level != level):
                            
                            # Actualizar código existente
                            existing.authorization_code_name = name
                            existing.authorization_level = level
                            existing.cucm_synced = True
                            existing.updated_at = datetime.utcnow()
                            
                            # Registrar auditoría
                            audit_entry = FacAudit(
                                authorization_code=code,
                                action="import_update",
                                admin_user=admin_username,
                                details=f"Código actualizado desde CUCM",
                                success=True
                            )
                            db.add(audit_entry)
                            updated_count += 1
                        else:
                            # No necesita actualización
                            existing.cucm_synced = True
                            unchanged_count += 1
                    else:
                        # Crear nuevo código
                        new_fac = FacCode(
                            authorization_code=code,
                            authorization_code_name=name,
                            authorization_level=level,
                            description=f"Importado desde CUCM el {datetime.utcnow().strftime('%Y-%m-%d %H:%M')}",
                            active=True,
                            cucm_synced=True
                        )
                        
                        db.add(new_fac)
                        
                        # Registrar auditoría
                        audit_entry = FacAudit(
                            authorization_code=code,
                            action="import_create",
                            admin_user=admin_username,
                            details=f"Código importado desde CUCM",
                            success=True
                        )
                        db.add(audit_entry)
                        created_count += 1
            
            # Registrar resumen
            summary_entry = FacAudit(
                authorization_code="SUMMARY",
                action="import_complete",
                admin_user=admin_username,
                details=f"Importación completada. Creados: {created_count}, Actualizados: {updated_count}, Sin cambios: {unchanged_count}",
                success=True
            )
            db.add(summary_entry)
            
            db.commit()
            print(f"Importación completada. Creados: {created_count}, Actualizados: {updated_count}, Sin cambios: {unchanged_count}")
        
    except Exception as e:
        # Registrar error
        audit_entry = FacAudit(
            authorization_code="N/A",
            action="import",
            admin_user=admin_username,
            details=f"Error de importación: {str(e)}",
            success=False
        )
        db.add(audit_entry)
        db.commit()
        print(f"Error importando códigos FAC desde CUCM: {str(e)}")
    
    finally:
        db.close()
